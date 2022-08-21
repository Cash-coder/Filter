import json

# files
CRAWLER_OUTPUT_PHONEMAD =  r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\phone_m_output.json'
FILTER_T2_OUTPUT        =  r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\FILTER_T2_OUTPUT.xlsx" 

# crawler json data names
ATTR1        = 'query_attribute_1'
DESCRIPTION  = 'description'
TARGET_CATEG = 'target_category'
QUERY_MODEL  = 'query_model'
PROD_URL     = 'prod_url'
PROD_ID      = 'complete_id'
TITLE        = 'title'
PRICE        = 'price'
PICS         = 'pics'
QUERY        = 'query'

# filter output cols
TARGET_CATEGORY_COL = 1
QUERY_COL          = 2
TITLE_COL          = 3
ATTR1_COL          = 7
# ATTR2_COL          = 8
PROD_URL_COL       = 9
WP_PRICE_COL       = 10
SUPPLIER_PRICE_COL = 11
WARRANTY_COL       = 15
SPECS_COL          = 19
PROD_ID_COL        = 18
QUERY_MODEL_COL    = 20
EDITED_PICS_COL    = 24
WP_SHORT_DESCRIPTION = 25
PROD_STATE_COL     = 26




def clean_excel(EXCEL_FILE):
    from openpyxl import load_workbook
    import logging

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    #starting at 3, delete all rows
    ws.delete_rows(3, ws.max_row+1)
    wb.save(EXCEL_FILE)

    logging.info(f'cleaned set_prod_db  and gaps_file to begin fresh writing')

def format_pics(pics):
    
    pics_string = ''
    for pic in pics:
        pics_string += pic
    
    return pics_string

def get_prod_state(description):

    from bs4 import BeautifulSoup as bs

    description = str(description)
    soup = bs(description, 'html.parser')

    text = soup.get_text()

    print(f'this is description: {text}')

    if 'Correcto' in text:
        prod_state = 'Estado correcto: 100% funcional, limpio y con accesorios.'

    # if 'estado Usado' in description:
    #     prod_state = 'El estado del artículo es usado, podría mostrar signos de uso'
    # elif 'estado Buen estado' in description:
    #     prod_state = 'El artículo se encuentra en buen estado'
    # elif 'estado Perfecto estado' in description:
    #     prod_state = 'El artículo se encuentra en perfecto estado'
    # elif 'estado A estrenar' in description:
    #     prod_state = 'El artículo se encuentra en perfecto estado, a estrenar.'
    
    return prod_state

def write_to_excel(data_to_dump):
    from openpyxl import  load_workbook

    wb = load_workbook(FILTER_T2_OUTPUT)
    ws = wb.active 

    wp_short_description = data_to_dump.get('wp_short_description') 
    target_category = data_to_dump.get('target_category') 
    query_model = data_to_dump.get('query_model') 
    prod_url    = data_to_dump.get('prod_url') 
    edited_pics = data_to_dump.get('edited_pics') 
    prod_state  = data_to_dump.get('prod_state') 
    prod_id = data_to_dump.get('cash_id') 
    attr1 = data_to_dump.get('attr1') 
    title = data_to_dump.get('title') 
    specs = data_to_dump.get('specs') 
    price = data_to_dump.get('price') 
    query = data_to_dump.get('query') 
    wp_price = data_to_dump.get('wp_price') 
    
    last_row = ws.max_row + 1

    # write warranty always 2 years
    ws.cell(row=last_row, column= WARRANTY_COL ,value=  '1 año') 

    ws.cell(row=last_row, column= QUERY_MODEL_COL ,value=  query_model ) 
    ws.cell(row=last_row, column= PROD_URL_COL ,value=  prod_url ) 
    ws.cell(row=last_row, column= EDITED_PICS_COL ,value=  edited_pics ) 
    ws.cell(row=last_row, column= PROD_STATE_COL ,value=  prod_state ) 
    ws.cell(row=last_row, column= ATTR1_COL ,value=  attr1 ) 
    ws.cell(row=last_row, column= TITLE_COL ,value=  title ) 
    ws.cell(row=last_row, column= SPECS_COL ,value=  specs ) 
    ws.cell(row=last_row, column= SUPPLIER_PRICE_COL ,value=  price ) 
    ws.cell(row=last_row, column= QUERY_COL ,value=  query ) 
    ws.cell(row=last_row, column= WP_PRICE_COL,value=  wp_price ) 
    ws.cell(row=last_row, column= WP_SHORT_DESCRIPTION,value=  wp_short_description ) 
    ws.cell(row=last_row, column= PROD_ID_COL ,value= prod_id ) 
    ws.cell(row=last_row, column= TARGET_CATEGORY_COL,value=  target_category) 

    wb.save(FILTER_T2_OUTPUT)

def get_specs(specs_html_string):
    from bs4 import BeautifulSoup as bs

    specs_text = ''
    soup = bs(specs_html_string, features='lxml')
    for li in soup.find_all('li'):
        label = li.find('span', class_="label").text
        value = li.find('span', class_="value").text.replace('\n','')
        
        spec = f'{label}{value}\n'
        specs_text += spec
    return specs_text

def apply_profit_margin(price, target_cat):
    import random 

    price = price.split(',')[0]
    price = int(price)
    
    if target_cat == 'videogames':
        profit = 4
        wp_price = price + profit
    
    else: 
        profit = 12
        wp_price = price + profit
    
    # declare taxes and fees
    taxes = profit * 0.21 #21%
    stripe_fee_rate  = 0.014 # 1.4%

    # taxes
    wp_price += taxes

    # calculate stripe fee
    stripe_fee = stripe_fee_rate * wp_price

    # add fee
    wp_price += stripe_fee

    wp_price = int(wp_price)

    # add attractive termination to the price: 55 -> 54,45
    terminator_options = [0.14,0.23,0.24,0.34,0.49,0.57,0.83,0.97]
    terminator = random.choice(terminator_options)

    final_price_decorated = wp_price + terminator

    # print(f'{wp_price} - {final_price_decorated}')
    return final_price_decorated


def run():
    
    # copy_move_file(selected_output_file, LOGS_FOLDER)

    #delete old entries to start fresh
    clean_excel(FILTER_T2_OUTPUT)


    with open(CRAWLER_OUTPUT_PHONEMAD, encoding='utf8') as json_file:
        scrapper_data = json.load(json_file)

        for i, item in enumerate(scrapper_data):        

            #before filtering by price, filter defective products with very cheap price that skews prices with very low prices from defective prods
            target_category = item[TARGET_CATEG]
            description     = item[DESCRIPTION]
            query_model     = item[QUERY_MODEL]
            prod_url    = item[PROD_URL]
            cash_id     = item[PROD_ID]
            attr1   = item[ATTR1]
            title   = item[TITLE]
            price   = item[PRICE]
            pics    = item[PICS]
            query = f'{query_model} {attr1}'

            # prod_state  = get_prod_state(description)
            # combine state and specs

            edited_pics = format_pics(pics)

            # 1.200 -> 1200
            price = price.replace('.','')

            wp_price    = apply_profit_margin(price, target_category)
            wp_short_description = 'Este artículo disfruta de una garantía de 1 año.\nPuedes probarlo durante 15 días.\nIncluye caja, cargador, funda de regalo y protector de pantalla.\nEnvío rápido 72 horas.'

            data_to_dump = {
                # 'prod_state':prod_state,
                'query_model':query_model,
                'wp_short_description':wp_short_description,
                'prod_url':prod_url,
                'wp_price':wp_price,
                'attr1':attr1,
                'title':title,
                'price':price,
                'query':query,
                'edited_pics':edited_pics,
                'cash_id':cash_id,
                'target_category':target_category
            }
            # {print(i) for i in data_to_dump}

            # items_list.append(data_to_dump)
            write_to_excel(data_to_dump)
            # print(f'processed {i+1}')

if __name__ == '__main__':
    run()