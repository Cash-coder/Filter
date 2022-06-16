# python filter resume_index channel
# python filter.py -    -c2
# python filter.py -455 -c2
# arg1 = resume_index in case of crush restart at index x
# arg2 = c2 = channel_2 use input and output files of channel 2

import filter_median_price
import time
import json
from os import extsep
from re import T
import traceback
from ebaysdk.finding import Connection
from ebaysdk.shopping import Connection as Shopping
from bs4 import BeautifulSoup
import logging
from openpyxl import load_workbook
from authentications import DEEPL_AUTH_KEY

# py functions files
import funcs_currency



# read crawler json output 

# USING FUNCS FROM (filter_median_lower_price_funcs.py), BUT IMPORTED AS FUNCTIONS AND EXECUTED HERE
# get list prods_filtered_by_median_low

# for prod in scrapper:
    # if prod_price < median_low for that model + attr: continue
    # get all prod data
    # apply funcs (wp_price, pics, detect warranty, color, GBP/USD to EUR  ...)
        # chek pics_db, if we have pics for that prod, use those pics
    # filter (price, payment methods, vendor votes, ...)
    # write to output file FILTER_OUTPUT.xlsx



logging.basicConfig(level=logging.INFO)

WEB_PICS_DB    = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\WEB_PICS_DB.xlsx"
INPUT_FILE     = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\crawler_output.json"
INPUT_FILE_C2  = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\crawler_output_C2.json"
OUTPUT_FILE    = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\FILTER_OUTPUT.xlsx" 
OUTPUT_FILE_C2 = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\FILTER_OUTPUT_CHANNEL2.xlsx" 
LOGS_FOLDER    = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\logs_folder"

# WARNING!! THIS FILE PATH HAS ALSO TO BE SET/UPDATED IN funcs_currency.py
CURRENCY_EQUIVALENCES_FILE = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\filter\exchange_to_EUR.txt'


# JSON FILTER INPUT NAMES, CRAWLER OUTPUT
#NOW USING CSV, IN DESUSE
EBAY_QUERY_NAME     = 'query'
EBAY_TITLE_NAME     = 'title'
EBAY_PRICE_NAME     = 'price'
EBAY_SHIPPING_TIME  = 'shipping_time'
EBAY_VARIABLE_PROD_NAME = 'variable_prod'
EBAY_RETURNS_NAME   = 'returns'
EBAY_SHIPPING_PRICE = 'shipping_price'
EBAY_ID_NAME        = 'ebay_article_id'
EBAY_PROD_URL_NAME  = 'prod_url'
EBAY_VENDOR_NAME    = 'ebay_vendor'
EBAY_SELLER_VOTES_NAME='seller_votes'
# EBAY_CATEGORY_NAME  = 'category'
EBAY_PAYMENT_NAME   = 'payment_methods'
EBAY_PROD_SPECS_NAME= 'prod_specs'
EBAY_PROD_STATE_NAME= 'product_state'
EBAY_PROD_DESCRIPTION_NAME = 'prod_description'
EBAY_SERVED_AREA_NAME = 'served_area'
EBAY_REVIEWS_NAME     = 'reviews'
EBAY_PROD_SOLD_OUT_NAME = 'product_sold_out_text'
EBAY_IMPORT_TAXES_NAME  = 'import_taxes'
TARGET_CATEGORY_NAME    = 'target_category' #used later in importer, if smartphone, import to smartphones
TARGET_ATTR_1_NAME   = 'query_attribute_1'
TARGET_ATTR_2_NAME   = 'query_attribute_2'
TARGET_MODEL_NAME    = 'query_model'
TARGET_PROD_STATE    = 'query_prod_state'
EBAY_PICS_URLS       = 'ebay_pics'
EBAY_SUBTITLE        = 'subtitle'
EBAY_IFRAME          = 'iframe_description_url'
SUBTITLE_NAME        = 'subtitle'

# #CSV ROWS N, FILTER INPUT, CRAWLER OUTPUT
# EBAY_QUERY_NAME       = 2
# EBAY_TITLE_NAME       = 0
# EBAY_PRICE_NAME       = 1
# EBAY_SHIPPING_TIME    = 3
# EBAY_VARIABLE_PROD_NAME=4
# EBAY_RETURNS_NAME     = 5
# EBAY_SHIPPING_PRICE   = 6
# EBAY_ID_NAME          = 7
# EBAY_PROD_URL_NAME    = 8
# EBAY_VENDOR_NAME      = 9
# EBAY_SELLER_VOTES_NAME= 10
# EBAY_CATEGORY_NAME    = 11
# EBAY_PAYMENT_NAME     = 12
# EBAY_PROD_SPECS_NAME  = 13
# EBAY_PROD_STATE_NAME  = 14
# # EBAY_PROD_DESCRIPTION_NAME=
# EBAY_SERVED_AREA_NAME  = 15
# EBAY_REVIEWS_NAME      = 16
# EBAY_PROD_SOLD_OUT_NAME= 17
# EBAY_IMPORT_TAXES_NAME= 18
# TARGET_CATEGORY_NAME  = 19
# TARGET_ATTR_1_NAME    = 20
# TARGET_ATTR_2_NAME    = 21
# TARGET_MODEL_NAME     = 22
# TARGET_PROD_STATE     = 23
# EBAY_PICS_URLS        = 24
# COLORS_AVAILABLE      = 25   

#FILTER_OUTPUT COLUMNS
QUERY_COL             = 1
EBAY_TITLE_COL        = 2
SUBTITLE_COL          = 3
EBAY_PROD_DESCRIPTION_COL = 4
EBAY_PROD_URL_COL     = 5
DETECTED_COLOR_COL    = 6
DETECTED_WARRANTY_COL = 7
EBAY_TOTAL_PRICE_COL  = 8
# CHECKMARK RESERVED  = 9 RESERVED!
# EBAY_VENDOR_NOTES_COL = 10 # NOT USED
TARGET_ATTR_1_COL     = 10 
TARGET_PROD_STATE_COL = 11
# EBAY_PROD_DESCRIPTION_LINK = 7 #description already included in prod_url
EBAY_PROD_STATE_COL   = 12
#EMPTY CELL TO LEAVE THE ACCEPT MARK= 10
EBAY_SELLER_VOTES_COL = 13
EBAY_RETURNS_COL      = 14
EBAY_SHIPPING_TIME_COL= 15
EBAY_SHIPPING_PRICE_COL=16
#from here I don't care for the order, random
EBAY_PRICE_COL     = 17
WP_PRICE_COL       = 18
PICTURES_COL       = 19
# PROD_BRAND_COL   = 17
EBAY_PROD_ID_COL   = 20
EBAY_CATEGORY_COL  = 21
TARGET_CATEGORY_COL= 22
                #   =23 #  NOT USED, AVAILABLE
TARGET_ATTR_2_COL   =24 
EBAY_PROD_SPECS_COL =25
EBAY_VENDOR_NAME_COL=26
WP_SHIPPING_TIME_COL=27
#                  = 28
MODEL_COL           =29


# performance metrics
def print_performance(n_prods, n_accepted_items):
    import time
    seconds = time.perf_counter()
    minuts  = seconds / 60
    minuts  = int(minuts)

    percentage_accepted_items = n_accepted_items / n_prods * 100
    percentage_accepted_items = str(percentage_accepted_items).split('.')[1][:2]

    print(f'it took {minuts} minuts to filter all {n_prods} units')
    print(f'accepted {n_accepted_items} of {n_prods} | {percentage_accepted_items}% of pass rate')
    
def apply_wp_price(price, target_category):
    import random

    # add benefit based on category
    benefit = get_benefit(target_category)
    price += benefit

    # add taxes
    tax_rate = benefit * 0.21 #21%
    price += tax_rate

    # add stripe fee
    stripe_fee_rate = 0.014 # 1.4%
    stripe_fee = stripe_fee_rate * price
    price += stripe_fee

    #add attractive termination to the price: 55 -> 54,45
    terminator_options = [0.14,0.23,0.24,0.34,0.49,0.57,0.83,0.97]
    terminator = random.choice(terminator_options)
    final_price_decorated = price + terminator
    #delete unwanted decimals
    final_price_decorated = round(final_price_decorated,2)
    # print(ebay_total_price,'\t', 'f_price_deco',final_price_decorated, '\t',"margin:",margin,'\t','benefit', benefit)
    # print(f'input_price:{input_price} output_price: {final_price_decorated} \t benefit: {benefit}, taxes {tax_rate}, stripe_fee_n {stripe_fee}')

    return final_price_decorated

# apply benefit margin based on category. 12€ smartphones, 4€ videogames
def get_benefit(target_category):
    # videogames doesn't exist yet, only consoles and videogames, we have to split it
    if target_category == 'videogames':
        benefit = 4
    else:
        benefit = 12
    return benefit

# now using a simpler cheaper margin version
# def apply_wp_price(ebay_total_price):
    
#     if int(ebay_total_price) >= 1000:
#         margin = 0.025
#     elif int(ebay_total_price) in range(800,1000):
#         margin = 0.030 # 3.5%  
#     elif int(ebay_total_price) in range(600,800):
#         margin = 0.035 # 4%  
#     elif int(ebay_total_price) in range(400,600):
#         margin = 0.045 # 5%  
#     elif int(ebay_total_price) in range(300,400):
#         margin = 0.055 # 6%  
#     elif int(ebay_total_price) in range(200,300):
#         margin = 0.065 # 8%  
#     elif int(ebay_total_price) in range(100,200):
#         margin = 0.10 # 10%  
#     elif int(ebay_total_price) in range(50,100):
#         margin = 0.12 # 12%  
#     elif int(ebay_total_price) in range(20,50):
#         margin = 0.30 # 30%  
#     elif int(ebay_total_price) in range(0,20):
#         margin = 0.5 # 50%

#     stripe_fee = 0.014 # 1.4%
#     benefit = ebay_total_price * margin
#     taxes = benefit * 0.21 #21%
    
#     #not 100% correct, stripe charges in the final amount
#     stripe_fee_n =  (ebay_total_price + benefit + taxes) * stripe_fee
#     final_price = ebay_total_price + benefit + taxes + stripe_fee_n
    
#     #apply the coupon code used in the ads
#     coupon_code_discount = 0.05 #5%
#     final_price = final_price * coupon_code_discount + final_price

#     #add attractive termination to the price: 55 -> 54,45
#     terminator_options = [0.14,0.23,0.24,0.34,0.49,0.57,0.83,0.97]
#     terminator = random.choice(terminator_options)
#     final_price_decorated = final_price + terminator
#     #delete unwanted decimals
#     final_price_decorated = round(final_price_decorated,2)
#     # print(ebay_total_price,'\t', 'f_price_deco',final_price_decorated, '\t',"margin:",margin,'\t','benefit', benefit)
#     print(f'ebay_total_price: {ebay_total_price} \t benefit: {benefit}, taxes {taxes}, stripe_fee_n {stripe_fee_n}, finalP: {final_price_decorated} ')

#     return final_price_decorated

#ebay API, in desuse
def get_ebay_pictures(ebay_prod_id):
    #not working because of ebay authentication issue
    try:
        api = Shopping(config_file='ebay.yaml')
        request = {'ItemID':ebay_prod_id}

        response = api.execute('GetSingleItem',request)

        pictures_url_list = []

        soup = BeautifulSoup(response.content,'lxml')
        #print(soup.prettify())
        pictures_url = soup.find_all('pictureurl')
        for url in pictures_url:
            url = str(url)
            url = url.replace('<pictureurl>','')
            url = url.replace('</pictureurl>','')
            pictures_url_list.append(url)
            print(url)
        return pictures_url_list
    except Exception as e:
        print(e)
        pass

#replace the last url parametter for s-l600.jpg to get the big pic instead of the thumbnail
def make_ebay_pics_urls(url_list):

    max_n_pics = 6

    def change_tail(url):
        chunks = url.split('/')
        tail   = chunks[-1]
        modified_url = url.replace(tail, 's-l500.jpg')
        return modified_url
    
    def convert_to_str(lst):
        pics_string = ''
        for i, pic in enumerate(lst):
            # max n pics processed
            if i > max_n_pics:
                continue
            pic = str(pic)
            pics_string += pic
            pics_string += ','
        return pics_string

    # if string == only 1 url
    if isinstance(url_list, str): 
        modified_url = change_tail(url_list)
        return modified_url

    # if lsit == several urls
    if isinstance(url_list, list): 
        modified_urls = []
        for url in url_list:
            new_url = change_tail(url)
            modified_urls.append(new_url)
        
        #convert from list to string to write to xlsx
        #to avoid error can't convert to excel, convert from list to string
        pics_string = convert_to_str(modified_urls)

        return pics_string
    

#I have to update, now we've 2 db's
#in this case will be web_pics_db
def check_pics_db(target_model, target_attr):
    from openpyxl import load_workbook

    #for row in web_pics_file:
        #row_pic_name
        #row_pic_model
        # if name and model == target name and model:
            #append all pics from that row to list
            #return list
    #if at the end of file, the list is == 0, return 'there aren\'t any pics'
    #else, return the list

    wb = load_workbook(WEB_PICS_DB)
    ws = wb.active
    
    pics_list = []
    for row in ws.iter_rows(min_row=2):
        row_number = row[0].row

        pic_model   = ws.cell(row=row_number, column=1).value
        pic_attr    = ws.cell(row=row_number, column=2).value
        # print('----------', pic_model, pic_attr)

        if target_model == pic_model and target_attr == pic_attr:
            n = 3
            for _ in range(15): #15 pics as max
                pic_url = ws.cell(row=row_number, column=n).value
                if pic_url != None and pic_url != '':
                    pics_list.append(pic_url)
                    n += 1
                else:
                    break
        else:
            continue
    if len(pics_list) > 0:
        return pics_list
    else:
        return 'there aren\'t any pics in pics_db for this item'


# takes raw html and gives a number of days|and the original: "5|5 march to 10 march"
def get_wp_shipping_time(ebay_shipping_time):
    from bs4 import BeautifulSoup as bs

    try:
        ebay_shipping_time = str(ebay_shipping_time)
        soup = bs(ebay_shipping_time, features='lxml')
        shipping_bs_separated = soup.get_text(separator="///")

        splitted = shipping_bs_separated.split('///')

        try:
            days = f"from {splitted[3]} to {splitted[5]}"
        except IndexError:
            days = shipping_bs_separated.replace('///', ' ')
        # print(f"this is days: {days}")

        #can improve and get the number of days using below old func
        return days

    except Exception as e:
        print(f'exception in get_wp_shipping_time(): {e}\nebay_shipping_time:{ebay_shipping_time:}')
        traceback.print_exc

        
    ### old func, maybe useful later    
    # try:
    #     #split ebay dates "vie. 9 jul. y el lun. 12 jul."
    #     ebay_date = ebay_shipping_time.split('.')
    #     ##get today's current day and month number
    #     current_day = str(date.today()).split('-')[2]
    #     current_month = str(date.today()).split('-')[1]
    #     current_month = int(current_month)
    #     current_month_name = calendar.month_name[current_month].lower()
    #     current_month_letters = current_month_name[0:3]

    #     # it can be one date(jul 12) or a range between 2 days (jul. 12 and jul. 16)
    #     #if it's a range of 2 dates:
    #     if len(ebay_date) == 5: 
    #         #select chunk like (9 jul)
    #         ebay_first_date = ebay_date[1]
    #         #select the number (9)
    #         ebay_first_day = ebay_first_date.split(' ')[1]
    #         #select the month name (jul)
    #         ebay_first_month = ebay_first_date.split(' ')[2]
    #         #second_date = ebay_date[3]
    #         #second_day = second_date.split(' ')[1]
    #         #ebay_second_month = second_date.split(' ')[2]
            
    #         #first letters from current month name to compare with ebay's
    #         # if it's the same month: jul == jul        
    #         if current_month_letters == ebay_first_month:
    #             shipping_days = int(ebay_first_day) - int(current_day)

    #             if shipping_days == 1:
    #                 wp_shipping_text = "Envío en 24h"
    #             elif shipping_days == 2:
    #                 wp_shipping_text = "Envío en 48h"
    #             elif shipping_days == 3:
    #                 wp_shipping_text = "Envío en 72h"
    #             elif shipping_days > 3: #if thre's more than 3 days return just the number of days
    #                 return shipping_days
    #             #UPDATE RETURN ONLY THE DAYS
    #             # return wp_shipping_text, shipping_days
    #             return shipping_days

    #         #if it's a different month, like jul-aug
    #         elif current_month_letters != ebay_first_month:
    #             #sample date: today's 27 jul, shipping arrival on 2 aug = 6 days
    #             # (from today's number to the end month) + ebay 1º day
    #             # (31-27)+2
    #             now = datetime.datetime.now()
    #             current_month_total_days = calendar.monthrange(now.year,now.month)[1]        
    #             daysto_end_month = current_month_total_days - now.month
    #             wp_shipping_days = daysto_end_month + int(ebay_first_day)
    #             return wp_shipping_days
    
    #     elif len(ebay_date) == 3: #there's only one date, like in "vie. 12 jul."
    #         ebay_day_number = ebay_shipping_time.split('.')[1]
    #         ebay_day_number = ebay_day_number.split(' ')[1]
    #         ebay_month_name = ebay_shipping_time.split('.')[1].split(' ')[2]

    #         #if it's the same month
    #         if current_month_letters == ebay_month_name:
    #             shipping_days = int(ebay_day_number) - int(current_day)
    #             if shipping_days == 1:
    #                 wp_shipping_text = "Envío en 24h"
    #             elif shipping_days == 2:
    #                 wp_shipping_text = "Envío en 48h"
    #             elif shipping_days == 3:
    #                 wp_shipping_text = "Envío en 72h"
    #             elif shipping_days > 3: #if thre's more than 3 days return just the number of days
    #                 return shipping_days
    #             return shipping_days
            
    #         #if it's a different month
    #         elif current_month_letters != ebay_month_name:

    #             now = datetime.datetime.now()
    #             current_month_total_days = calendar.monthrange(now.year,now.month)[1]        
    #             daysto_end_month = current_month_total_days - now.month
    #             shipping_days = daysto_end_month + int(ebay_day_number)

    #             if shipping_days == 1:
    #                 wp_shipping_text = "Envío en 24h"
    #             elif shipping_days == 2:
    #                 wp_shipping_text = "Envío en 48h"
    #             elif shipping_days == 3:
    #                 wp_shipping_text = "Envío en 72h"
    #             elif shipping_days > 3: #if thre's more than 3 days return just the number of days
    #                 return shipping_days

    #             #UPDATE RETURN ONLY SHIPPING DAYS
    #             # return wp_shipping_text, shipping_days
    #             return shipping_days
    # except Exception as e:
    #     print('exception in get_wp_shipping_time()',e)

# def apply_prod_brand(ebay_title):
    # #do the same, using data from target
    # if 'iphone' or 'Iphone' or 'iPhone' or 'IPhone' in ebay_title:
        # prod_brand = 'iphone'
    # elif 'samsung' or 'Samsung' in ebay_title:
        # prod_brand = 'samsung'
    # return prod_brand

def copy_move_file(src_dir, dst_dir, mode='time'):
    '''file_name, src_dir, dst_dir, mode=time or no_time // absolute paths'''
    #if mode=time -> include time on file title, elif normal -> not include time
    import os
    import shutil
    import datetime

    #detect file_name and file_format
    file = src_dir.split('\\')[-1]
    file_name = file.split('.')[0]
    file_format = file.split('.')[1]
    file_format = '.' + file_format

    #create time id to rename the file
    now = str(datetime.datetime.now())[:16]
    now = now.replace(' ', '_')
    now = now.replace(':', '-')

    #include time or not based on specified mode
    if mode == 'time':
        abs_path = file_name + str(now) + file_format
    elif mode == 'no_time':
        abs_path = file_name + file_format

    dst_path = dst_dir + '\\' + abs_path

    shutil.copy(src_dir, dst_path)
    print(f'file {file} moved to {dst_path}')

def clean_excel(EXCEL_FILE):
    from openpyxl import load_workbook
    import logging

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    #starting at 3, delete all rows
    ws.delete_rows(3, ws.max_row+1)
    wb.save(EXCEL_FILE)

    logging.info(f'cleaned set_prod_db  and gaps_file to begin fresh writing')

def color_detector(ebay_title):
    # print(f'ebay_title: {ebay_title}')
    ebay_title = ebay_title.lower()
    # ordering all colors in the same order, the using index, if the first match in iatlina, you know it's negro, second mattch= azul
    spanish_colors = ['negro',  'azul',  'marrón',  'gris',  'verde',   'naranja',  ' rosa',    'violeta',  'rojo', 'blanco',    'amarillo', 'oro', 'plata']
    english_colors = ['black',  'blue',  'brown',   'grey',   'green',   'orange',   'pink',    'purple',   'red',  'white',    'yellow',   'gold', 'silver']
    italian_colors = ['nero',   'blu',   'marrone', 'grigio', 'verde',  'arancione', 'rosa',    'viola',    'rosso', 'bianco',   'giallo'   'oro',  'argento'] 
    german_colors  = ['schwarz', 'blau', 'braun',    'grau',  'grün',   'orange',    'rosa',     'lila',     'rot',   'weiß',    'gelb',     'gold', 'silber']
    french_colors  = [ 'noir',    'bleu', 'brun',   'gris',   'vert',   'orange',   'rose',      'pourpre', 'rouge', 'blanc',   'jaune',   '----',  'argent'] #mising golden in french because it's 'or', will make a lot of false postives

    #some standalone colors
    if 'grafito' in ebay_title:
        return 'grafito'

    #spanish
    for color in spanish_colors:
        if color in ebay_title:
            detected_color = color
            return detected_color
    #english
    for color in english_colors:
        if color in ebay_title:
            i = english_colors.index(color)
            detected_color = spanish_colors[i]
            return detected_color
    #italian
    for color in italian_colors:
        if color in ebay_title:
            i = italian_colors.index(color)
            detected_color = spanish_colors[i]
            return detected_color
    #german
    for color in german_colors:
        if color in ebay_title:
            i = german_colors.index(color)
            detected_color = spanish_colors[i]
            return detected_color
    #french
    for color in french_colors:
        if color in ebay_title:
            i = french_colors.index(color)
            detected_color = spanish_colors[i]
            return detected_color

    return None

#get warranty searching for text match
def detect_warranty(subtitle, description):
    #notice monate != monaten
    one_year  = ['garantía 1 año', '1 año garantía','garantía 12 meses', '12 meses garantía', '12 meses de garantía', '12 monate gewährleistung', '12 monaten gewährleistung' ,'12 monate herstellergarantie', '1 jahr gewährleistung' , '1 jahr herstellergarantie', '12 mesi garanzia', 'garanzia 12 mesi' , 'garanzia 1 anni']
    half_year = ['6 meses garantía', 'garantía 6 meses', '6 monate gewährleistung' , '6 monate herstellergarantie', '6 mesi garanzia', 'garanzia 6 mesi']
    two_years = ['2 años de garantía','24 meses garantía','garantía 24 meses' ,'24 monate gewährleistung' , '24 monate herstellergarantie', '2 jahr gewährleistung' , '2 jahr herstellergarantie', '24 mesi garanzia', 'garanzia 24 mesi' , 'garanzia 1 anni']

    #check in subtitle
    try:
        subtitle = str(subtitle)
        subtitle = subtitle.lower()
        for item in one_year:
            if item in subtitle:
                return '1 año'
        for item in half_year:
            if item in subtitle:
                return '6 meses'
        for item in two_years:
            if item in subtitle:
                return '2 años'
    except AttributeError as e:
        print('asdadasdsd')
        print(e)
        pass #subtitle not present

    #check in description
    try:
        description = str(description)
        description = description.lower()
        for item in one_year:
            if item in description:
                return '1 año'
        for item in half_year:
            if item in description:
                return '6 meses'
        for item in two_years:
            if item in description:
                return '2 años'
    except AttributeError: pass


def nlp_translate(text, target_language='es'):
    import requests
    import authentications #py file

    translate_api_key = authentications.RapidAPI_Key

    url = "https://nlp-translation.p.rapidapi.com/v1/translate"

    # protected words that won't translate
    protected_words = 'apple'
    
    # querystring = {"text":t,"to":"es","from":"de"}
    querystring = {"text":text, "to":target_language, "protected_words": protected_words}

    headers = {
        "X-RapidAPI-Host": "nlp-translation.p.rapidapi.com",
        "X-RapidAPI-Key": translate_api_key
    }

    response = requests.request("GET", url, headers=headers, params=querystring).json()
    # print(response)
    
    if response['translated_characters'] > 0:
        translated_text = response['translated_text']['es']
        
        try:
            origin_language = response['from']
            return translated_text, origin_language
        except KeyError: # when text is alredy in spanish doesn0t appear 'from'
            return translated_text, '---'
    else: # if no translation performed
        return text, '---'

def translate_specs(specs):
    
    # remove some parts we don't want, like "El articulo puede mostrar un deterioro"
    # a part of the text is spanish already
    # if you put into the translator it interprets it like spanish, so it won't translate
    # split the not_spanish part with \n
    # join in a string
    # translate
    # combine again translated with spanish
    # add the phrase "Texto traducido automáticamente desde el Alemán:"

    specs = specs.replace(' El artículo puede mostrar un deterioro ... ', '').replace('(en caso de ... ','')

    phrase_to_add_german  = "Texto traducido automáticamente del Alemán:\n\n"
    phrase_to_add_italian = "Texto traducido automáticamente del Italiano:\n\n"
    phrase_to_add_english = "Texto traducido automáticamente del Inglés:\n\n"
    phrase_to_add_french  = "Texto traducido automáticamente del Francés:\n\n"
    
    if 'Estado: ' in specs:
        splitted = specs.split('\n')
        text_to_translate    = splitted[2:]
        already_spanish_text = splitted[:1]

        joined_to_translate = "\n".join(text_to_translate)
        joined_spanish      = "\n".join(already_spanish_text)

        r = nlp_translate(joined_to_translate)
    else:
        r = nlp_translate(specs)

    translated, origin_lan = r
    # print(translated) 

    # avoid error 'joined_spanish' doesn't exist
    if 'joined_spanish' not in locals():
        joined_spanish = '' 

    if origin_lan   == 'en':
        combined = phrase_to_add_english + joined_spanish + "\n" + translated
    elif origin_lan == 'fr':
        combined = phrase_to_add_french + joined_spanish + "\n" + translated
    elif origin_lan == 'de' or origin_lan == 'no':
        combined = phrase_to_add_german + joined_spanish + "\n" + translated
    elif origin_lan == 'it':
        combined = phrase_to_add_italian + joined_spanish + "\n" + translated
    
    else: #for spanish 'es' and other languages
        combined = joined_spanish + "\n" + translated
    
    return combined


def get_ebay_vendor_notes(ebay_prod_specs):
    #vendor notes are OFTEN in the 2º place of the list prod_specs. Get that, if "Notas del vendedor in list", translate and return
    try:
        vendor_notes     = ebay_prod_specs.split('\n')[0]
        translated_notes = nlp_translate(vendor_notes)
        return translated_notes

        # for item in ebay_prod_specs:
        #     print(f'this is the item {item}')
        #     notes_title = item[0]
        #     vendor_notes = item[1]

        #     print(notes_title, vendor_notes, 'yyy')
        #     if notes_title == 'Notas del vendedor:':
        #         print('match!')
        #         translated_notes=deepl_translate(vendor_notes)
        #         print('translated_notes= ',translated_notes)
        #         return translated_notes
        
        # notes_title = raw_vendor_notes[0]
        # vendor_notes = raw_vendor_notes[1]
        
    except Exception as e:
        print('exception in get_ebay_vendor_notes',e)
        return None


def write_to_excel(data_to_dump, selected_output_file):

    wb = load_workbook(selected_output_file)
    ws = wb.active 

    query =         data_to_dump.get('query') 
    ebay_title=     data_to_dump.get('ebay_title')     
    prod_state =    data_to_dump.get('prod_state')
    target_category=data_to_dump.get('target_category')
    ebay_price =    data_to_dump.get('ebay_price')
    ebay_prod_id =  data_to_dump.get('ebay_prod_id')
    ebay_category = data_to_dump.get('ebay_category')
    ebay_prod_description =     data_to_dump.get('ebay_prod_description')
    target_attr_2 = data_to_dump.get('target_attr_2')
    ebay_price =    data_to_dump.get('ebay_price')
    ebay_shipping_price = data_to_dump.get('ebay_shipping_price')
    ebay_returns =  data_to_dump.get('ebay_returns')
    ebay_prod_url = data_to_dump.get('ebay_prod_url')
    ebay_prod_specs=data_to_dump.get('ebay_prod_specs')
    wp_price =      data_to_dump.get('wp_price')
    wp_shipping_time=data_to_dump.get('wp_shipping_time')
    target_attr_1 = data_to_dump.get('target_attr_1')
    ebay_pics     = data_to_dump.get('pictures')
    ebay_total_price=data_to_dump.get('ebay_total_price')
    # to compare prices more easily write an integer
    # ebay_vendor_notes=data_to_dump.get('ebay_vendor_notes')
    target_prod_state=data_to_dump.get('target_prod_state')
    detected_color=data_to_dump.get('detected_color')
    warranty    =data_to_dump.get('warranty')
    subtitle  =data_to_dump.get('subtitle')
    target_model  =data_to_dump.get('target_model')
    seller_votes  =data_to_dump.get('seller_votes')
    # prod_db_category=data_to_dump.get('prod_db_category')
    # ebay_shipping_time = data_to_dump.get('ebay_shipping_time') #using wp_shipping time instead
    # prod_brand =    data_to_dump.get('prod_brand')

    last_row = ws.max_row + 1

    ws.cell(row=last_row, column= QUERY_COL,value=  query)
    ws.cell(row=last_row, column= EBAY_TITLE_COL,value=     ebay_title)
    ws.cell(row=last_row, column= EBAY_PROD_STATE_COL,value=prod_state)
    ws.cell(row=last_row, column= EBAY_PRICE_COL,value=     ebay_price)
    ws.cell(row=last_row, column= EBAY_PROD_ID_COL,value=   ebay_prod_id)
    ws.cell(row=last_row, column= EBAY_CATEGORY_COL,value=  ebay_category)
    ws.cell(row=last_row, column= EBAY_PROD_DESCRIPTION_COL,value= ebay_prod_description)
    ws.cell(row=last_row, column= PICTURES_COL,value=        ebay_pics)
    ws.cell(row=last_row, column= TARGET_CATEGORY_COL,value= target_category)
    ws.cell(row=last_row, column= TARGET_ATTR_2_COL,value=  target_attr_2)
    ws.cell(row=last_row, column= EBAY_PRICE_COL,value=     ebay_price)
    ws.cell(row=last_row, column= EBAY_SHIPPING_PRICE_COL,value=ebay_shipping_price)
    ws.cell(row=last_row, column= EBAY_RETURNS_COL,value=   ebay_returns)
    ws.cell(row=last_row, column= EBAY_PROD_URL_COL,value=  ebay_prod_url)
    ws.cell(row=last_row, column= EBAY_PROD_SPECS_COL,value=ebay_prod_specs)
    ws.cell(row=last_row, column= WP_PRICE_COL,value=       wp_price)
    ws.cell(row=last_row, column= WP_SHIPPING_TIME_COL,value=  wp_shipping_time)
    ws.cell(row=last_row, column= TARGET_ATTR_1_COL,value=     target_attr_1)
    ws.cell(row=last_row, column= EBAY_TOTAL_PRICE_COL,value=  ebay_total_price)
    # ws.cell(row=last_row, column= EBAY_VENDOR_NOTES_COL,value= ebay_vendor_notes)
    ws.cell(row=last_row, column= TARGET_PROD_STATE_COL,value= target_prod_state)
    ws.cell(row=last_row, column= EBAY_SELLER_VOTES_COL,value= seller_votes)
    ws.cell(row=last_row, column= DETECTED_COLOR_COL ,value= detected_color)
    ws.cell(row=last_row, column= DETECTED_WARRANTY_COL ,value= warranty)
    ws.cell(row=last_row, column= SUBTITLE_COL ,value= subtitle)
    ws.cell(row=last_row, column= MODEL_COL ,value= target_model)
    # ws.cell(row=last_row, column= EBAY_SHIPPING_TIME_COL,value=ebay_shipping_time)

    wb.save(selected_output_file)
    # print(f'written to excel item {ebay_title}')


def get_textfrom_html(ebay_prod_description):
    from bs4 import BeautifulSoup
    try:
        text = BeautifulSoup(str(ebay_prod_description), features="html.parser").get_text()
        return text
    except Exception as e:
        print(f'in get_textfrom_html(): {e}')
        traceback.print_exc()

# delete text, convert from str to int
def process_shipping_price(ebay_shipping_price):
   
    if ebay_shipping_price == []:
        ebay_shipping_price = 0
        return ebay_shipping_price

    try:
        # used to debug
        original = ebay_shipping_price 

        if 'GRATIS' in ebay_shipping_price or 'gratis' in ebay_shipping_price:
            ebay_shipping_price = 0
            return ebay_shipping_price
        
        try:
            ebay_shipping_price = ebay_shipping_price.replace('EUR','').replace('GBP','').replace('USD','').replace(',','.').replace('aprox.','').replace('(','').replace(')','').replace('de gastos de envío','').strip()
            ebay_shipping_price = float(ebay_shipping_price)
        except Exception as e:
            print(f'Exception with ebay shipping price: original: {original}\nprocessed: <{ebay_shipping_price}>\nError: {e}')
            return 'error processing shipping price'
        
        try:
            if 'Rápido y gratis' in ebay_shipping_price:
                ebay_shipping_price = 0                
        except Exception as e:
            #print(e)
            pass
        
        try:
            if 'gratis' in ebay_shipping_price.lower():
                ebay_shipping_price = 0
        except:
            pass
            
            
        return ebay_shipping_price
    
    except Exception as e:
        print(e)
        traceback.print_exc
        return 'error processing shipping price'

#clean price and convert to float
def process_price(ebay_price):

    try:
        #in the case of a price with comma like "1.102,95 EUR"
        if ('.') in ebay_price and (',') in ebay_price: 
            # print("detected!!!", ebay_price)
            ebay_price = ebay_price.split(',')[0]
            ebay_price = ebay_price.replace('.','')
            ebay_price = float(ebay_price)

        else: #if its a number like  "516,95 EUR"
            ebay_price = ebay_price.replace('EUR','').replace(',','.').replace('c/u','').replace(' ','').strip()
            ebay_price = float(ebay_price)

        return ebay_price

    except Exception as e:
        print(f'error in process_price(): {e}')
        return 'price error'

def get_subtitle(raw_subtitle):
    
    if raw_subtitle:
        subtitle, origin_lan = nlp_translate(raw_subtitle)
    else:
        subtitle = ''

    return subtitle

def get_seller_votes(seller_votes):
    if seller_votes:
        try:
            seller_votes = int(seller_votes)
        except Exception as e:
            print(f'exception in int(seller_votes): {e}')
            traceback.print_exc()
            seller_votes = 100 #to get this prod pass the filter
    else: #if for some reason it didn't get seller_votes, put it in 100
        seller_votes = 100 #above of the limit to pass the votes filter 
    
    return seller_votes


def get_prod_description(ebay_prod_description):
    #sometimes descriptio is short, only text and helpful. Other times is an iframe full of stuff. Ignore when full
    if len(ebay_prod_description) > 1200:
        ebay_prod_description = 'Too long'
    else:
        ebay_prod_description = get_textfrom_html(ebay_prod_description)
        ebay_prod_description = ebay_prod_description.replace('\\t', '').replace(r'\\n', '')
    
    return ebay_prod_description

# if the seller is known for having spanish text, avoid translation = faster filter
def check_if_spanish_seller(ebay_vendor_name):
    identified_spanish_sellers = [
        'cashconverters_es'
    ]

    if ebay_vendor_name in identified_spanish_sellers:
        return True
    else:
         return False

def get_args():
    import sys
    # first arg  resume filtering in case of crush, specify as a parameter when calling from terminal
    # you have to manually save the output file because this will remove all old entries. So it will start from index x, but all processed data will be removed
    # or you can find the file in the logs folder
    
    # arg1 = resume_index in case of crush restart at index x
    try:
        first_arg = sys.argv[1]
        if first_arg != '-':
            resume_from_index_n = int(first_arg)
        else:
            resume_from_index_n = 0
    except IndexError: # if not specified start from the beginning
        resume_from_index_n = 0

    # arg2 = c2 = channel_2 use input and output files of channel 2
    try:
        second_arg = sys.argv[2]
        if second_arg == '-c2':
            selected_input_file  = INPUT_FILE_C2
            selected_output_file = OUTPUT_FILE_C2
    except IndexError:
        selected_input_file  = INPUT_FILE
        selected_output_file = OUTPUT_FILE
    
    #print selecteds to check by the user
    input_folder1  = selected_input_file.split('\\')[-1]
    output_folder2 = selected_output_file.split('\\')[-1]
    print(f'\nselected_input_file: {input_folder1}\nselected_output_file: {output_folder2}\n resume_from_index_n: {resume_from_index_n}\n')

    return selected_input_file, selected_output_file, resume_from_index_n

def clean_title(ebay_title):
    
    title = ebay_title.lower()
    
    # items to delete
    terms_to_remove = ['sin cara id', 'manzana','iva incl.', 'incl. iva', '(gsm)', 'gsm', 'at&t', 'at&amp;t', 'feria', 'cdma', '()', '/c 397']
    title = remove_from_title(terms_to_remove, title)

    #excelente estado
    l = ['a-ware', 'a (grado)', 'grado a']
    alt_phrase = 'estado excelente'
    title = substitute(l, title, alt_phrase)
    
    # muy buen estado
    l = ['muy bien', 'muy bueno', 'bien']
    alt_phrase = 'muy buen estado'
    title = substitute(l, title, alt_phrase)
    
    #estado aceptable
    l = ['b-ware', 'grado b', 'grado a/b', ' b ', 'akzeptabel']
    alt_phrase = 'estado aceptable'
    title = substitute(l, title, alt_phrase)
    
    # desbloqueado
    l = ['sin simlock', 'sin sim-lock', 'désimlocké', 'sin carrier', 'sin bloqueo de sim', 'sin bloqueo sim']
    alt_phrase = 'desbloqueado'
    title = substitute(l, title, alt_phrase)
    
    return title

# given title, change unwanted phrase for alternative phrase
def substitute(items_list, old_title, alternative_phrase):
    try:
        for item in items_list:
            if item in old_title:
                new_title = old_title.replace(item, alternative_phrase)
                return new_title
        return old_title
    except Exception as e:
        return old_title
        
def remove_from_title(items_list, title):
    
    for item in items_list:
        if item in title:
            title = title.replace(item, '')
    
    return title

def set_logger():
    
    # // Logging levels: Higher to lower
    # TRACE
    # DEBUG
    # INFO
    # WARN
    # ERROR
    # FATAL
    # OFF
    
    import logging
    import sys

    logger = logging.getLogger()
    logger.setLevel(logging.WARN)
    # formatter = logging.Formatter('%(asctime)s | %(levelname)s | %(message)s', 
                                # '%m-%d-%Y %H:%M:%S')
    formatter = logging.Formatter('%(message)s')

    stdout_handler = logging.StreamHandler(sys.stdout)
    stdout_handler.setLevel(logging.WARN)
    stdout_handler.setFormatter(formatter)

    file_handler = logging.FileHandler('logs.log')
    file_handler.setLevel(logging.WARN)
    file_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(stdout_handler)

    #in code:
    # import logging
    # logging.warning('Watch out!')


def process_import_taxes(ebay_import_taxes):
    # ebay_import_taxes = ebay_import_taxes.replace('£','').replace('US $','')
    if '£' in ebay_import_taxes:
        amount = ebay_import_taxes.replace('£','')
        ebay_import_taxes = funcs_currency.convert_amount_toEUR(amount, 'GBP')
    elif 'US $' in ebay_import_taxes:
        amount = ebay_import_taxes.replace('US $','')
        ebay_import_taxes = funcs_currency.convert_amount_toEUR(amount, 'USD')
    else:
        ebay_import_taxes = 0
    
    return ebay_import_taxes


### End definitions ###


def run():

    set_logger()

    # taking data from parameters passed when called the program
    selected_input_file, \
    selected_output_file, \
    resume_from_index_n = get_args()

    #make a copy in logs_folder
    copy_move_file(selected_output_file, LOGS_FOLDER)
    #delete old entries to start fresh
    clean_excel(selected_output_file)


    #for item in data:
        #check if item is broken,
        # check if price < low_median + tolerance
        # apply rest of the filter
        # if item seller votes < x
        #other filter rules...
        #items that pass filter:
            #try to use pics from pics_db, if no pics for that prod, use ebay's pics
            #get_wp_price
            #get_shipping_time
            #...
            #append to list
        #write from list to excel

    
    filtered_list = []
    # changed from median low to median high: now if product price is below median high we accept it.
    # median_low_price_list = filter_median_lower_price.run()
    median_low_price_list = filter_median_price.run(selected_input_file)
    # [print(item) for item in median_low_price_list]

    n_accepted_items = 0
    with open(selected_input_file, encoding='utf8') as json_file:
        scrapper_data = json.load(json_file)

        print(f'starting process from index: {resume_from_index_n}')
        for item in scrapper_data[resume_from_index_n:]:        

            try:

                # here resume_index is used to print the current position of the process: item 5 of 10
                print(f'\nprod processed: {resume_from_index_n} of {len(scrapper_data)} ')
                resume_from_index_n += 1

                #before filtering by price, filter defective products with very cheap price that skews prices with very low prices from defective prods
                ebay_title  =  item[EBAY_TITLE_NAME]
                prod_state  =  item[EBAY_PROD_STATE_NAME]
                subtitle    =  item[SUBTITLE_NAME]
                ebay_url    =  item[EBAY_PROD_URL_NAME]
                item_description = item[EBAY_PROD_DESCRIPTION_NAME]
                item_description = "\n".join(item_description)
                
                # is_defective =  filter_median_price.check_defective_prod(ebay_title, prod_state)
                is_defective =  filter_median_price.check_defective_prod(ebay_title, prod_state, item_description, subtitle)
                #IMPROVEMENT in prod description I could mark when "faulty", "craced", "smashed" ... are present in prod description or title, and like colors, include a column with detected faults
                
                if is_defective:
                    print(f'faulty item {ebay_title, prod_state, ebay_url}')
                    continue
                
                # check median_lower price + tolerance
                # take the only 2 needed data to check price filter median low
                target_model  = item[TARGET_MODEL_NAME]
                target_attr_1 = item[TARGET_ATTR_1_NAME]
                
                # avoid filter searching for "None" string in titles
                # if target_attr_1 == None:
                    # target_attr_1 = ''
                if target_model == None:
                    target_model = ''

                # filter total price using median_low
                model_and_attribute = target_model + f' {target_attr_1}'
                ebay_price =      item[EBAY_PRICE_NAME]

                
                # avoid prods from australia
                if 'AUD' in ebay_price:
                    continue
                
                # this puts the price to 0, to correct you have to search in console the str: "error in ebay shipping_price ebay_id" 
                ebay_shipping_price = item[EBAY_SHIPPING_PRICE]
                ebay_shipping_price = process_shipping_price(ebay_shipping_price)
                if ebay_shipping_price == 'error processing shipping price':
                    print(f'error in ebay shipping_price ebay_id: {ebay_prod_id}')
                    ebay_shipping_price = 0
                
                #if error price to 0, search in console Error in ebay_price, product with ebay_id:
                ebay_price = process_price(ebay_price)
                if ebay_price == 'price error':
                    print(f'Error in ebay_price, product with ebay_id: { ebay_prod_id}')
                    ebay_price = 0
                
                ebay_import_taxes  =item[EBAY_IMPORT_TAXES_NAME]
                # process prices taxes, shipping, $€ exchange
                if ebay_import_taxes:
                    ebay_import_taxes = process_import_taxes(ebay_import_taxes)
                else:
                    ebay_import_taxes = 0
                
                ebay_total_price = ebay_price + ebay_shipping_price + ebay_import_taxes

                # r_median_high_filter = filter_median_price.filter_median_high_price(model_and_attribute, ebay_total_price, median_high_price_list)
                r_median_low_filter = filter_median_price.filter_median_high_price(model_and_attribute, ebay_total_price, median_low_price_list)
                logging.warning(f'model_attr:{model_and_attribute}-title {ebay_title}')
                # if doesn't pass the filter continue
                if not r_median_low_filter:
                    logging.warning(f'---NOT ACCEPTED')
                    continue

                # apply the rest of the filter
                variable_prod =   item[EBAY_VARIABLE_PROD_NAME]
                # remove some common kws like AT&T
                ebay_title = clean_title(ebay_title)
                # ebay_reviews       =item[EBAY_REVIEWS_NAME]
                # ebay_category =   item[EBAY_CATEGORY_NAME]
                payment_methods = item[EBAY_PAYMENT_NAME]
                sold_out_text =   str(item[EBAY_PROD_SOLD_OUT_NAME])
                area_served =     item[EBAY_SERVED_AREA_NAME]
                target_category = item[TARGET_CATEGORY_NAME]
                target_attr_2 =   item[TARGET_ATTR_2_NAME]
                ebay_pics     =   item[EBAY_PICS_URLS]
                ebay_vendor_name   =item[EBAY_VENDOR_NAME]
                target_prod_state  =item[TARGET_PROD_STATE]
                query =           item[EBAY_QUERY_NAME]
                ebay_shipping_time = item[EBAY_SHIPPING_TIME]
                ebay_returns =    item[EBAY_RETURNS_NAME]
                ebay_prod_id =    item[EBAY_ID_NAME]
                ebay_prod_url=    item[EBAY_PROD_URL_NAME]
                ebay_subtitle   = item[EBAY_SUBTITLE]
                # ebay_iframe_url = item[EBAY_IFRAME]

                wp_price = apply_wp_price(ebay_total_price, target_category)

                raw_subtitle = item[SUBTITLE_NAME]
                subtitle = get_subtitle(raw_subtitle)
                
                raw_seller_votes = item[EBAY_SELLER_VOTES_NAME]
                seller_votes = get_seller_votes(raw_seller_votes)

                ebay_prod_specs = item[EBAY_PROD_SPECS_NAME]            
                
                #i.e: cash converters is always in spanish, no need to translate, identify more spanish sellers to filter faster
                spanish_seller = check_if_spanish_seller(ebay_vendor_name)
                if not spanish_seller:
                    ebay_prod_specs        = translate_specs(ebay_prod_specs)
                    ebay_title, origin_lan = nlp_translate(ebay_title)

                #IMPROVEMENT in prod description I could mark when "faulty", "craced", "smashed" ... are present in prod description or title, and like colors, include a column with detected faults
                raw_ebay_prod_description = item[EBAY_PROD_DESCRIPTION_NAME]
                ebay_prod_description  = get_prod_description(raw_ebay_prod_description)
                warranty               = detect_warranty(ebay_subtitle, ebay_prod_description)
                #detected_faults = detect_faults()
                #write faults column

                ebay_vendor_notes= get_ebay_vendor_notes(ebay_prod_specs)
                # detected_color   = color_detector(ebay_title)
                
                if area_served == None: area_served='not result'
                
                #cash converters includes color in specs
                # if ebay_vendor_name == 'cashconverters_es':
                    # detected_color = color_detector(ebay_prod_specs)
                
                #if item is variable write to sheet2
                if variable_prod != None: #avoid product if it's a variable prod
                    print("this item is variable",item['title'])
                    # write to another file or sheet ?
                    # write_to_excel(data_to_dump, OUTPUT_FILE, 'variables')
                    continue
                elif '[]' not in sold_out_text : # if the product is NOT sold out it's an empty list
                    print('prod sold out',item['title'])
                    continue
                elif  seller_votes < 30: #if very little sells
                    print(f'not enough votes, current votes: {seller_votes}')
                    continue
                elif 'PayPal' not in payment_methods or 'Visa' not in payment_methods:
                    print('not payment',item['title'], f'payment_methods: {payment_methods} \n')
                    continue
                elif ebay_price == '':
                    print('no price',item['title'])
                    continue
                elif ebay_shipping_price == '':
                    print('not shipping price ',item['title'])
                    continue
                elif ebay_shipping_price == 'local pick up':
                    print(f'this prod has local pick up, hence no shipping: {ebay_prod_url}')
                    continue
                elif 'Solo recogida local' in area_served:
                    print('only local pick up no shipping prod: ',item['title'])
                    continue
                
        ###################### THIS GOES IN SCRAPPER#################

                #check if there're pictures for this prod in pics_db
                # pictures = check_pics_db(target_model, target_attr_2)
                # if any pic in pics_db, use ebay's pictures
                # if pictures == 'there aren\'t any pics in pics_db for this item':
                    # logging.info(f'there aren\'t any pics in pics_db for this item <{target_model} {target_attr_2}>')
                    # print(f'going to search this ebay_id {ebay_prod_id}')
                pictures = make_ebay_pics_urls(ebay_pics)
                    # pictures = get_ebay_pictures(ebay_prod_id)# through api
                
                wp_shipping_time = get_wp_shipping_time(ebay_shipping_time) 
                
                #apply our category based on Ebay's category
                # prod_db_category = apply_category(ebay_category)
                # prod_brand = apply_prod_brand(ebay_title)

                #print("there are products with wanted characteristics!!")
                data_to_dump = {
                    'query':query,
                    'ebay_title':ebay_title,
                    'prod_state':prod_state,
                    # 'prod_db_category':prod_db_category,
                    # 'brand_prod_filter':prod_brand,
                    'ebay_price':ebay_price,
                    'ebay_shipping_price':ebay_shipping_price,
                    # 'ebay_shipping_time':ebay_shipping_time,
                    'ebay_returns':ebay_returns,
                    'ebay_prod_id':ebay_prod_id,
                    'ebay_prod_url':ebay_prod_url,
                    # 'ebay_category':ebay_category,
                    'ebay_prod_specs':ebay_prod_specs,
                    'ebay_prod_description':ebay_prod_description,
                    'wp_price':wp_price,
                    'pictures':pictures,
                    'wp_shipping_time':wp_shipping_time,
                    'target_category':target_category,
                    'target_attr_1':target_attr_1,
                    'target_attr_2':target_attr_2,
                    'ebay_total_price':ebay_total_price,
                    'ebay_vendor_notes':ebay_vendor_notes,
                    'target_prod_state':target_prod_state,
                    'seller_votes':seller_votes,
                    # 'detected_color':detected_color,
                    'warranty':warranty,
                    'target_model':target_model,
                    'subtitle':subtitle,
                }

                filtered_list.append(data_to_dump)
                #variables I don't need in filtered csv
                #'payment_methods':payment_methods, sold_out_text,'seller_votes':seller_votes,
                #area_served,'var_prod':variable_prod

                # with open('filter_output.json','w',) as new_file: #,encoding='utf8')
                    # json.dump(entry, new_file, indent=4)
                    # # new_file.write(",")
                    # # new_file.close()

                write_to_excel(data_to_dump, selected_output_file)
                n_accepted_items += 1

            except Exception as e:
                print(e)
                continue

        n_all_items = len(scrapper_data)
        print_performance(n_all_items, n_accepted_items)


if __name__ == '__main__':
    run()

    ################################
    #csv.field_size_limit(sys.maxint)
    # maxInt = sys.maxsize

    # #this is to avoid error while reding row "row character lentgh exceeds maximum"
    # while True:
    #     # decrease the maxInt value by factor 10 
    #     # as long as the OverflowError occurs.

    #     try:
    #         csv.field_size_limit(maxInt)
    #         break
    #     except OverflowError:
    #         maxInt = int(maxInt/10)
    ######################
