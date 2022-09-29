import json
from pstats import SortKey
import traceback

#settings
max_pics = 3

# files
CRAWLER_OUTPUT_CASHCON = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\cash_crawler_output.json'
FILTER_T2_OUTPUT =  r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\FILTER_T2_OUTPUT.xlsx" 

# crawler json data names
ATTR1        = 'query_attribute_1'
DESCRIPTION  = 'description_text'
TARGET_CATEG = 'target_category'
QUERY_MODEL  = 'query_model'
PROD_URL     = 'prod_url'
CASH_ID      = 'prod_id'
TITLE        = 'title'
SPECS        = 'specs'
PRICE        = 'price'
PICS         = 'pics'
QUERY        = 'query'

# filter T2 output cols 
CATEGORY_T2          = 1
COMPLETE_QUERY_T2    = 2
WP_TITLE_T2          = 3
ATTR1_T2             = 7
ATTR2_T2             = 8
SUPPLIER_PROD_URL_T2 = 9
WP_PRICE_T2          = 10
SUPPLIER_PRICE_T2    = 11
SHIPPING_TIME_T2     = 13
SHIPPING_PRICE_T2    = 14
WARRANTY_T2          = 15
RETURNS_T2           = 16
SPECS_T2             = 19
SUPPLIER_PROD_ID_T2  = 18
QUERY_MODEL_T2       = 20
SUPPLIER_PICS_URL_T2 = 24
WP_SHORT_DESCRIPTION_T2 = 25
PROD_STATE_T2           = 26

# filter output cols
CATEGORY_T2 = 1
COMPLETE_QUERY_T2       = 2
WP_TITLE_T2       = 3
ATTR1_T2       = 7
ATTR2_T2       = 8
SUPPLIER_PROD_URL_T2    = 9
WP_PRICE_T2    = 10
SUPPLIER_PRICE_T2  = 11
WARRANTY_T2    = 15
SPECS_T2       = 19
SUPPLIER_PROD_ID_T2    = 18
QUERY_MODEL_T2 = 20
SUPPLIER_PICS_URL_T2 = 24
WP_SHORT_DESCRIPTION = 25
PROD_STATE_T2  = 26



def clean_excel(EXCEL_FILE):
    from openpyxl import load_workbook
    import logging

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    #starting at 3, delete all rows
    ws.delete_rows(3, ws.max_row+1)
    wb.save(EXCEL_FILE)

    logging.info(f'cleaned set_prod_db  and gaps_file to begin fresh writing')


def edit_pic_urls(pics):
    
    pics_string = ''

    for pic in pics[::max_pics]:
        pics_string += pic + ','
    
    return pics_string

def get_prod_state(description):
    description = description.lower()

    if 'estado usado' in description:
        prod_state = 'El ha sido usado con anterioridad, podría mostrar signos de uso'
    elif 'estado buen estado' in description:
        prod_state = 'El artículo se encuentra en Buen Estado'
    elif 'estado perfecto estado' in description:
        prod_state = 'El artículo se encuentra en Perfecto Estado'
    elif 'estado excelente' in description:
        prod_state = 'El artículo se encuentra en Estado Excelente'
    elif 'estado a estrenar' in description:
        prod_state = 'El artículo se encuentra a Estrenar.'
    
    return prod_state

def write_to_excel(data_to_dump):
    from openpyxl import  load_workbook

    wb = load_workbook(FILTER_T2_OUTPUT)
    ws = wb.active 

    wp_short_description = data_to_dump.get('wp_short_description') 
    target_category = data_to_dump.get('target_category') 
    query_model = data_to_dump.get('query_model') 
    prod_url    = data_to_dump.get('prod_url') 
    edited_pics = data_to_dump.get('edited_pics') 
    prod_state  = data_to_dump.get('prod_state') 
    cash_id = data_to_dump.get('cash_id') 
    attr1 = data_to_dump.get('attr1') 
    title = data_to_dump.get('title') 
    specs = data_to_dump.get('specs') 
    price = data_to_dump.get('price') 
    query = data_to_dump.get('query') 
    wp_price = data_to_dump.get('wp_price') 
    
    last_row = ws.max_row + 1

    # write warranty always 2 years
    ws.cell(row=last_row, column= WARRANTY_T2 ,value=  '2 años') 

    ws.cell(row=last_row, column= QUERY_MODEL_T2 ,value=  query_model ) 
    ws.cell(row=last_row, column= SUPPLIER_PROD_URL_T2 ,value=  prod_url ) 
    ws.cell(row=last_row, column= SUPPLIER_PICS_URL_T2 ,value=  edited_pics ) 
    ws.cell(row=last_row, column= PROD_STATE_T2 ,value=  prod_state ) 
    ws.cell(row=last_row, column= ATTR1_T2 ,value=  attr1 ) 
    ws.cell(row=last_row, column= WP_TITLE_T2 ,value=  title ) 
    ws.cell(row=last_row, column= SPECS_T2 ,value=  specs ) 
    ws.cell(row=last_row, column= SUPPLIER_PRICE_T2 ,value=  price ) 
    ws.cell(row=last_row, column= COMPLETE_QUERY_T2 ,value=  query ) 
    ws.cell(row=last_row, column= WP_PRICE_T2,value=  wp_price ) 
    ws.cell(row=last_row, column= WP_SHORT_DESCRIPTION,value=  wp_short_description ) 
    ws.cell(row=last_row, column= SUPPLIER_PROD_ID_T2,value=  cash_id ) 
    ws.cell(row=last_row, column= CATEGORY_T2,value=  target_category) 

    wb.save(FILTER_T2_OUTPUT)

def get_specs(specs_html_string):
    from bs4 import BeautifulSoup as bs

    specs_text = ''
    soup = bs(specs_html_string, features='lxml')
    for li in soup.find_all('li'):
        # label = li.find('span', class_="label").text
        # value = li.find('span', class_="value").text.replace('\n','')
        # spec = f'{label}{value}\n'
        spec = li.text.replace('\n', '')
        spec += '\n'
        specs_text += spec
        # print(specs_text)

    return specs_text

def apply_profit_margin(price, target_cat):
    import random 

    price = int(price)
    
    if target_cat == 'videogames':
        profit = 4
        wp_price = price + profit
    
    else: 
        profit = 12
        wp_price = price + profit
    
    # declare taxes and fees
    taxes = profit * 0.21 #21%
    stripe_fee_rate  = 0.014 # 1.4%

    # taxes
    wp_price += taxes

    # calculate stripe fee
    stripe_fee = stripe_fee_rate * wp_price

    # add fee
    wp_price += stripe_fee

    wp_price = int(wp_price)

    # add attractive termination to the price: 55 -> 54,45
    terminator_options = [0.14,0.23,0.24,0.34,0.49,0.57,0.83,0.97]
    terminator = random.choice(terminator_options)

    # if 500 transform in 499
    if str(wp_price)[-1] == '0': # if the last digit == 0:
        wp_price -= 1

    # print(f' this is wp price: {wp_price}')
    final_price_decorated = wp_price + terminator

    # print(f'{wp_price} - {final_price_decorated}')
    return final_price_decorated

def WriteFromFilterT2ToDb():
    from openpyxl import load_workbook

    wbr = load_workbook(FILTER_T2_OUTPUT)
    wsr = wbr.active
    n_uploaded_prods = 1
    total_rows       = wsr.max_row
    for row in wsr.iter_rows(min_row=3):
        row_n = row[0].row #get row number

        query= wsr.cell(row=row_n, column=COMPLETE_QUERY_T2).value

        #used to ignore blank excell cells
        if query == None: continue

        dataToInsert = (
                #category
                wsr.cell(row=row_n,column=CATEGORY_T2).value,
                #complete query
                wsr.cell(row=row_n, column=COMPLETE_QUERY_T2).value,
                # wpTitle
                wsr.cell(row=row_n,column=WP_TITLE_T2).value,
                #wooID
                n_uploaded_prods,
                #targetProdState
                wsr.cell(row=row_n,column=CATEGORY_T2).value,
                #sourceState
                wsr.cell(row=row_n,column=PROD_STATE_T2).value,
                #attr1
                wsr.cell(row=row_n,column=ATTR1_T2).value,
                #attr2
                wsr.cell(row=row_n,column=ATTR2_T2).value,
                #supplierTotalPrice
                wsr.cell(row=row_n,column=SUPPLIER_PRICE_T2).value,
                #wpPrice
                wsr.cell(row=row_n,column=WP_PRICE_T2).value,
                #wpUrl
                'https://someUrl',
                #supplierProdURL
                wsr.cell(row=row_n,column=SUPPLIER_PROD_URL_T2).value,
                #shippingTime
                wsr.cell(row=row_n,column=SHIPPING_TIME_T2).value,
                #shippingPrice
                wsr.cell(row=row_n,column=SHIPPING_PRICE_T2).value,
                #warrany
                wsr.cell(row=row_n,column=WARRANTY_T2).value,
                #returns
                wsr.cell(row=row_n,column=RETURNS_T2).value,
                #webPics
                wsr.cell(row=row_n,column=SUPPLIER_PICS_URL_T2).value,
                #webPicsId
                n_uploaded_prods,
                #supplierProdId
                wsr.cell(row=row_n,column=SUPPLIER_PROD_ID_T2).value,
                #sourceSpecs
                wsr.cell(row=row_n,column=SPECS_T2).value,
                #targetModel
                wsr.cell(row=row_n, column= QUERY_MODEL_T2).value,
                #adPics
                'ad pics names',
                #wpShortDescription
                wsr.cell(row=row_n, column= WP_SHORT_DESCRIPTION).value,
                #AmazonReviewsDone
                False,
                )

#return the 3 lowest prices for each product
def getLowerPricesItems(scrapperData):
    
    bestPrices = []
    nMaxItems = 2

    #multivariable sort, title and price 
    sortedList = sorted(scrapperData, key=lambda x: (x['title'], x['price']))

    #used to differentiate items in list
    currentTitle = sortedList[0]['title'] 
    flag = 0
    for item in sortedList:
        title = item.get('title')
        #used to print and debug
        # price = item.get('price')
        # price = item.get('price').replace(' €', '').replace('.', '').replace(',', '.')
        # price = float(price)

        #the list is sorted by price and title
        #if title is same and flag below max:
            #append, flag +1    
        #if title is same and flag 3 or more
            #continue
            #this way you only append nMax items
        #if title is new:
            #flag to 0
            #update title

        if title == currentTitle and flag <= nMaxItems:
            bestPrices.append(item)
            flag += 1
        elif title == currentTitle and flag > nMaxItems:
            continue
        elif title != currentTitle:
            flag = 0
            currentTitle = title
            # bestPrices.append(item)

    [print(item['title'], item['price']) for item in bestPrices]
    return bestPrices


def run():
    
    # copy_move_file(selected_output_file, LOGS_FOLDER)

    #delete old entries to start fresh
    clean_excel(FILTER_T2_OUTPUT)

    with open(CRAWLER_OUTPUT_CASHCON, encoding='utf8') as json_file:
        scrapper_data = json.load(json_file)

        getLowerPricesItems(scrapper_data)

        # for i, item in enumerate(scrapper_data):        

        #     try:

        #         #before filtering by price, filter defective products with very cheap price that skews prices with very low prices from defective prods
        #         target_category = item[TARGET_CATEG]
        #         description     = item[DESCRIPTION]
        #         query_model     = item[QUERY_MODEL]
        #         prod_url    = item[PROD_URL]
        #         cash_id     = item[CASH_ID]
        #         attr1   = item[ATTR1]
        #         title   = item[TITLE]
        #         specs   = item[SPECS]
        #         price   = item[PRICE]
        #         pics    = item[PICS]
        #         query   = item[QUERY]

        #         prod_state  = get_prod_state(description)
        #         specs_text  = get_specs(specs)  
        #         # combine state and specs
        #         specs_text = f'{prod_state}\n\n{specs_text}'

        #         edited_pics = edit_pic_urls(pics)

        #         # 1.200 -> 1200
        #         price = price.split(',')[0].strip()
        #         price = price.replace('.','')

        #         wp_price = apply_profit_margin(price, target_category)
        #         wp_short_description = 'Este artículo disfruta de una garantía de 2 años completos.\nPuedes probarlo durante 30 días.\nEnvío rápido 72 horas.'

        #         data_to_dump = {
        #             'query_model':query_model,
        #             'wp_short_description':wp_short_description,
        #             'prod_url':prod_url,
        #             'wp_price':wp_price,
        #             'attr1':attr1,
        #             'title':title,
        #             'specs':specs_text,
        #             'price':price,
        #             'query':query,
        #             'edited_pics':edited_pics,
        #             'prod_state':prod_state,
        #             'cash_id':cash_id,
        #             'target_category':target_category
        #             }

        #         # items_list.append(data_to_dump)
        #         write_to_excel(data_to_dump)
        #         print(f'processed {i+1}')

        #     except Exception as e:
        #         print(e)
        #         traceback.print_exc()


if __name__ == '__main__':
    run()