from openpyxl import load_workbook


# go row by row
# if find bad_stuff_sign
    # remove row


INPUT_FILE  = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\FILTER_OUTPUT.xlsx" 

bad_stuff_signs = [
    'montaje', 'ensamblaje', 'para', 'reposamanos', 'reposamuñecas', 'pantalla completa', 'pieza original', 'conjunto de pantalla', 'conjunto pantalla', 'pantalla conjunto', 'conjunto', 'plata lcd completa',
    'barra superior', 'topcase', 'top case',
    'repuesto', 'pantalla rota', 'roto', 'pantalla rajada', 'pantalla dañada', 'daños', 'desmontaje',
    'placa madre', 'placa base', 'tarjeta madre', 'motherboard',  'logic card', 'placa lógica', 'defectuoso', 'defectuosa',
    ]


def _check(items_list, text):
    text = text.lower()

    for item in items_list:
        if item in text:
            print(f'found bad sign: {item} in text: {text}')
            return True


wb = load_workbook(INPUT_FILE)
ws = wb.active
    
for row in ws.iter_rows(min_row=2):
    row_number = row[0].row

    title = row[1].value
    title = str(title).lower()
    

    r = _check(bad_stuff_signs, title)
    if r: 
        ws.delete_rows(row[0].row, 1)

wb.save(INPUT_FILE)