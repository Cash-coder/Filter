from openpyxl import load_workbook

# reads web_pics_db sheet4, from row 2
#get all paths from row. Like https://...
#converto to [{"id": "https://..."}, {},]


FILTER_OUTPUT1   =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\FILTER_OUTPUT.xlsx"
FILTER_T2_OUTPUT =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\FILTER_T2_OUTPUT.xlsx"

# LOGS_FOLDER    =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\logs_folder"
WEB_PICS_DB    =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\WEB_PICS_DB.xlsx"
WEB_PICS_SHEET ='Sheet3'
ADS_PICS_DB    =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\ADS_PICS_DB.xlsx"
ADS_PICS_SHEET ='Sheet7' 

START_ROW = 3 #start reading file at row x

def get_pics_paths(row_number):
    wb = load_workbook(WEB_PICS_DB)
    ws = wb['Sheet4']

    pics_list = []
    n = 1
    for _ in range(12): #15 pics as max
        pic_url = ws.cell(row=row_number, column=n).value
        # print(pic_url)
        if pic_url != None and pic_url != '':
            pics_list.append(pic_url)
            n += 1
        else:
            break
    return pics_list

def format_webpics(pics_list):
    formatted_list = []
    for link in pics_list:
        if link == None: 
            print('link == None, continue')
            continue

        entry = {"id":f"{link}"}
        # print(entry)
        formatted_list.append(entry)
    return formatted_list


def clean_excel(EXCEL_FILE, sheet):
    from openpyxl import load_workbook
    import logging

    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet]

    #starting at 2, delete all rows
    ws.delete_rows(2, ws.max_row+1)
    wb.save(EXCEL_FILE)
    logging.info(f'cleaned file')


def run():
    wb = load_workbook(WEB_PICS_DB)
    ws = wb['Sheet4']

    for row in ws.iter_rows(min_row=2):
        rown = row[0].row

        pics_paths = get_pics_paths(rown)
        formatted_paths = format_webpics(pics_paths)

        print(formatted_paths)
    
    a = input('do you want to clean excel ? y/n ')
    if a == 'y':
        clean_excel(WEB_PICS_DB, 'Sheet4')


if __name__ == '__main__':
    run()

