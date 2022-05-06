import time
from openpyxl import load_workbook


file   = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\filter\test.xlsx"
trans  = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\filter\trans.xlsx"

wb = load_workbook(file)
ws = wb.active

wbt = load_workbook(trans)
wst = wbt.active



def nlp_translate(text, target_language='es'):
    import requests
    import authentications #py file

    translate_api_key = authentications.RapidAPI_Key

    url = "https://nlp-translation.p.rapidapi.com/v1/translate"

    # querystring = {"text":t,"to":"es","from":"de"}
    querystring = {"text":text, "to":target_language}

    headers = {
        "X-RapidAPI-Host": "nlp-translation.p.rapidapi.com",
        "X-RapidAPI-Key": translate_api_key
    }

    response = requests.request("GET", url, headers=headers, params=querystring).json()
    translated_text = response['translated_text']['es']

    return translated_text

def write_response(r):

    splitted = r.split('|-|') 
    # print(f'this is response {r} \n')
    
    for i in range(len(splitted)):
        start_row = wst.max_row + 1
        translated_chunk = splitted[i].replace('|-|', '')
        wst.cell(row=start_row, column=1, value= translated_chunk)
        # print(f'written {translated_chunk}')
    wbt.save(trans)


def serial_func():
    t1_start = time.time()

    row  = 1
    for cell in ws['A']:
        text = cell.value
        translated = nlp_translate(text)
        wst.cell(row= row, column=1, value= translated)
        row+= 1
    wbt.save(trans)

    print(f'time: {time.time() - t1_start}')


def combined_func():
    t1_start = time.time()

    combined_str = ""
    for cell in ws['S']:
        text = cell.value

        # translate before reaching 5k characters
        combined_length = len(text) + len(combined_str)
        if combined_length >= 5000:
            print(combined_length)
            r = nlp_translate(combined_str)
            combined_str = ''
            write_response(r)
        else:
            text += '|&|'
            combined_str += text
        
    print(f'time: {time.time() - t1_start}')


def combined_func2():
    t1_start = time.time()
    
    # a = ws.cell_range(min_col=19, max_col=19, min_row=3, max_row=6)
    
    # for row in ws.iter_rows(min_row=3):
    #     rown = row[0].row
    #     text = row[18].value
        # print(text)
    
    row = 2
    combined_str = ''
    for i in range(ws.max_row):
        row += 1
        current_text = ws.cell(row=row, column=19).value
        next_text    = ws.cell(row=row+1, column=19).value
        
        # print(f'{current_text}  \n{next_text}')

        # if current_len + next_len < 5000:
        # if current_len + next_len + combined_str_len < 5000:
        #     combined_str += current_text

        if next_text == None:
            next_text = ''
        if current_text == None:
            continue

        if len(current_text) + len(next_text) + len(combined_str) < 5000:
            current_text += '|&|'
            combined_str += current_text
            if i == ws.max_row - 2:
                r = nlp_translate(combined_str)
                write_response(r)
                combined_str = ''
        else:
            r = nlp_translate(combined_str)
            write_response(r)
            combined_str = ''

    print(f'time: {time.time() - t1_start}')
        

# 5x5
# if less than 5000 : append to comb_str
    # else: translate now
#translate
# next 5

def func():
    pass
    
def combined_func3():
    t1_start = time.time()

    comb_str = ''
    row = 2
    while True:
        for i in range(5):
            row += 1
            text = ws.cell(row=row, column=19).value
            
            if text == None:
                print(f'None cell in row {row}')
                return

            r = func()
            if len(text) + len(comb_str) < 5000:
                text += '|&|'
                comb_str += text
            else:
                r = nlp_translate(comb_str)
                resto = text

    print(f'time: {time.time() - t1_start}')


def combined_func4():

    resto = ''
    combined_str = ""
    for cell in ws['A']:
        text = cell.value

        if text == None: continue

        if resto != '':
            combined_str += resto
            resto = ''

        if len(text) + len(combined_str) < 5000:
            text += '|-|'
            combined_str += text
        else:
            print(f'len combined: {len(combined_str)}')
            r = nlp_translate(combined_str)
            write_response(r)
            combined_str = ''
            resto = text



# global variable to hold
#for each 5 cells
# for cell in 5_cells:
    #if len text + len

# serial_func()
# combined_func()

t1_start = time.time()
# combined_func4()
serial_func()
print(f'time: {time.time() - t1_start}')


