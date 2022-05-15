#si no tiene specs, copiar specs de item mismo modelo y attr1, attr2

import traceback
from openpyxl import load_workbook

#read filter t1 output, search for column with checkmark
#compare target_prod_state == ebay_prod_state, if not, ebay state wins
#put the web pics id's
    # search for item model + attr1
    # manually check if they're the right color, mark that prod as reviewed
    #if no web_pics and ebay_pics == weird_ delete that row
    # write to an excel
#same for ad pics
#move file to logs
#clean file for the next round


FILTER_OUTPUT1   =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\FILTER_OUTPUT.xlsx"
FILTER_T2_OUTPUT =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\FILTER_T2_OUTPUT.xlsx"

# LOGS_FOLDER    =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\logs_folder"
WEB_PICS_DB    =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\WEB_PICS_DB.xlsx"
WEB_PICS_SHEET ='Sheet4' #where the id's are
ADS_PICS_DB    =r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\ADS_PICS_DB.xlsx"
ADS_PICS_SHEET ='Sheet7' 

START_ROW = 3 #start reading file at row x

# FROM FILTER OUTPUT 1
QUERY_COL             = 1
TARGET_PROD_STATE_COL = 2
EBAY_PROD_STATE_COL   = 3
EBAY_TITLE_COL        = 4
AVAILABLE_COLORS_COL  = 5
DETECTED_COLOR_COL    = 6
DETECTED_WARRANTY_COL = 7
EBAY_TOTAL_PRICE_COL  = 8
CHECKMARK             = 9
# EBAY_VENDOR_NOTES_COL = 10
TARGET_ATTR_1_COL     = 10 
SUBTITLE_COL          = 11
# EBAY_PROD_DESCRIPTION_LINK = 7 #description already included in prod_url
EBAY_PROD_URL_COL     = 12
#EMPTY CELL TO LEAVE THE ACCEPT MARK= 10
EBAY_SELLER_VOTES_COL = 13
EBAY_RETURNS_COL      = 14
EBAY_SHIPPING_TIME_COL= 15
EBAY_SHIPPING_PRICE_COL=16
#from here I don't care for the order, random
EBAY_PRICE_COL     = 17
WP_PRICE_COL       = 18
PICTURES_COL       = 19
# PROD_BRAND_COL   = 17
EBAY_PROD_ID_COL   = 20
EBAY_CATEGORY_COL  = 21
TARGET_CATEGORY_COL= 22
                #   =23 #  NOT USED, AVAILABLE
# TARGET_ATTR_2_COL  = 24 #in phones now I don't target for colors
EBAY_PROD_SPECS_COL= 25
EBAY_VENDOR_NAME_COL=26
WP_SHIPPING_TIME_COL=27
EBAY_PROD_DESCRIPTION_COL = 28
MODEL_COL           =29

# FOR FILTER PUTPUT 2
TARGET_CATEGORY2_COL = 1
COMPLETE_QUERY2_COL  = 2
WP_TITLE2_COL        = 3
WOO_ID2_COL          = 4
TARGET_STATE2_COL    = 5
SOURCE_STATE2_COL    = 6
ATTR1_1_COL          = 7
ATTR1_2_COL          = 8
SUPPLIER_TOTAL_PRICE2_COL = 9
WP_PRICE2_COL        = 10
WP_URL2_COL          = 11
SUPPLIER_URL2_COL    = 12
EBAY_SHIPPING_TIME2_COL  = 13
EBAY_SHIPPING_PRICE2_COL = 14
WARRANTY2_COL        = 15
EBAY_RETURNS2_COL    = 16
WEB_PICS2_COL        = 17
EBAY_ID2_COL         = 18
SOURCE_SPECS         = 19
TARGET_MODEL2_COL    = 20
VARIABLE_PROD2_COL   = 21
ADS_PICS2_1_COL      = 22
ADS_PICS2_2_COL      = 23
EBAY_PICS2_COL       = 24
WP_SHORT_DESCRIPTION_COL = 25



def copy_move_file(src_dir, dst_dir, mode='time'):
    '''file_name, src_dir, dst_dir, mode=time or no_time // absolute paths'''
    #if mode=time -> include time on file title, elif normal -> not include time
    import os
    import shutil
    import datetime

    #detect file_name and file_format
    file = src_dir.split('\\')[-1]
    file_name = file.split('.')[0]
    file_format = file.split('.')[1]
    file_format = '.' + file_format

    #create time id to rename the file
    now = str(datetime.datetime.now())[:16]
    now = now.replace(' ', '_')
    now = now.replace(':', '-')

    #include time or not based on specified mode
    if mode == 'time':
        abs_path = file_name + str(now) + file_format
    elif mode == 'no_time':
        abs_path = file_name + file_format

    dst_path = dst_dir + '\\' + abs_path

    shutil.copy(src_dir, dst_path)
    print(f'file {file} moved to {dst_path}')

def clean_excel(EXCEL_FILE):
    from openpyxl import load_workbook
    import logging

    wb = load_workbook(EXCEL_FILE)
    ws1 = wb['Sheet1']
    ws2 = wb['Sheet2']

    #starting at 2, delete all rows
    ws1.delete_rows(3, ws1.max_row+1)
    wb.save(EXCEL_FILE)
    ws2.delete_rows(3, ws2.max_row+1)
    wb.save(EXCEL_FILE)

    logging.info(f'cleaned set_prod_db  and gaps_file to begin fresh writing')

# copy_move_file(FILTER_OUTPUT_PATH, LOGS_FOLDER)
# clean_excel(FILTER_OUTPUT_PATH)

not_found = 0
#get pics using model and color (target_attr)
def get_pics_ids(file, target_model, target_attr):
    from openpyxl import load_workbook
    global not_found

    #row by row, if item and attr1 = target, return pics, else, return not found
    #select between read web_pics_db or ads_pics_db
    if file == 'web_pics':
        WB = WEB_PICS_DB
        WS = WEB_PICS_SHEET
    elif file == 'ads_pics':
        WB = ADS_PICS_DB
        WS = ADS_PICS_SHEET  

    # print(file, WB, WS)
    # print(f'file:{file}, target_model:{target_model}, target_attr:{target_attr}')

    wb = load_workbook(WB)
    ws = wb[WS]
    
    all_pics  = [] 
    pics_list = [] # create to avoid Error when pic is not present ?
    for row in ws.iter_rows(min_row=2):
        row_number = row[0].row
        # print(f'this is row N {row_number}')

        #web_pics_db model and attr
        pic_model   = ws.cell(row=row_number, column=1).value
        pic_attr    = ws.cell(row=row_number, column=2).value
        pic_model = pic_model.strip() 
        pic_attr = pic_attr.strip() 

        if pic_model == None or pic_attr == None:
            continue

        #if model == target_model, color == color: get those pics // using in instead of == to allow variations in colors: negro,gris,grafito
        #THIS IS OLD FUNC, WHEN NEED TO MATCH MODEL AND COLOR, NEW FUNC NEED MATCH ONLY MODEL
        if target_model == pic_model and target_attr in pic_attr:
            if file == 'web_pics':
                pics_list = get_pics_by_row(file, row_number)
                if 'not found' in pics_list[0]:
                    continue
                for pic in pics_list:
                    all_pics.append(pic)
                print(f'2 len {len(pics_list)}')
            
            #func designed ony to walla, to get all ad pics
            elif file == 'ads_pics':
                pics_list = get_walla_ad_pics(target_model)
        else:
            # print(f'excel_pic_model: {pic_model}, excel_pic_attr: {pic_attr}')
            continue

    # if len(pics_list) > 0:
    if pics_list:
        print('final return pics_list')
        return pics_list
    else:
        not_found += 1
        print(f'not found this model <{target_model}> with this attribute <{target_attr}> not_found {not_found}')
        return 'not found'
    
def get_walla_ad_pics(target_model):
    
    wb = load_workbook(ADS_PICS_DB)
    ws = wb[ADS_PICS_SHEET]

    all_pics = []
    for row in ws.iter_rows():
        rown = row[0].row
        excel_model = row[0].value

        if excel_model == target_model:
            pic1 = ws.cell(row=rown, column=3).value
            pic2 = ws.cell(row=rown, column=4).value
            all_pics.append(pic1)
            all_pics.append(pic2)
            # pics = get_pics_by_row('ads_pics', rown)
            # for pic in pics:
                # all_pics.append(pics)
    return all_pics

    
#get the pics using the row n. mind that's the same func for web_pics and ads_pics, be carefull with changes
def get_pics_by_row(file, row_number):
    # print(f'row_number:{row_number}, type: {type(row_number)}')

    if file == 'web_pics':
        WB = WEB_PICS_DB
        WS = WEB_PICS_SHEET
    elif file == 'ads_pics':
        WB = ADS_PICS_DB
        WS = ADS_PICS_SHEET  

    wb = load_workbook(WB)
    ws = wb[WS]

    #for web_pics
    if file == 'web_pics':
        pics_list = []
        n = 3
        for _ in range(12): #15 pics as max
            pic_url = ws.cell(row=row_number, column=n).value
            if pic_url != None and pic_url != '':
                pics_list.append(pic_url)
                n += 1
            else:
                break
    #for ads_pics, I need to return a string with 2 paths, not a list, because later it makes a list of lists
    elif file == 'ads_pics':
        n = 3
        pics_list = ''
        for _ in range(12): #15 pics as max
            pic_url = ws.cell(row=row_number, column=n).value
            if pic_url != None or pic_url != '' or pic_url != ' ':
                #string instead of list
                pics_list = pics_list +','+ pic_url
                n += 1
            else:
                break        

    print(f'file {file}, row {row_number} founded {len(pics_list)} pics')

    if len(pics_list) > 0:
        print(f'pics_listn {len(pics_list)} returning pics_list {pics_list}')
        #web_pics need format [{"id":"link"}, ...]
        # FORMATTED IN WOO_API, HERE RETURN ID, ID, ID
        # if file == 'web_pics':
            # web_pics_formatted = format_webpics(pics_list)
            # return web_pics_formatted 
        return pics_list # return ads path
    else:
        return 'not found'

#NOT USED, FORMATTED IN WOO_API.PY
#web_pics need to convert from link to [{"id":"link"}, ...]
def format_webpics(pics_list):
    formatted_list = []
    for id in pics_list:
        entry = {"id":int(id)}
        # print(entry)
        formatted_list.append(entry)
    return formatted_list


#compare prods state between target_state and ebay_state. If diff then ebay_state wins. FI: target == new but ebay_state == usado. Then prod_state is usado, not new
#ebay known prod states: Usado, Nuevo, Abierto sin usar
def compare_states(target_prod_state, ebay_prod_state):
    if target_prod_state == 'new' and ebay_prod_state != 'Nuevo':
        prod_state = 'not_new'
    elif target_prod_state == 'not_new' and ebay_prod_state == 'Nuevo':
        prod_state = 'new'
    else: #if targee_state match ebay_state       
        prod_state = target_prod_state
    
    return prod_state

def get_short_description(ebay_returns, warranty):
    #get number of days from ebay_returns
    #if warranty, code it. If not, use alt text

    # split string by spaces, if chunk is digit, put it in list
    number_ofdays = [int(s) for s in ebay_returns.split(' ') if s.isdigit()]
    
    # Truth table: 4 combinations: days/warranty, days/no warranty, no days/warranty, no days/no warranty
    if len(number_ofdays) > 0:
        number_ofdays = str(number_ofdays[0]) #from list to string
        
        if warranty != None: # return and warranty
            short_des = f'Este artículo dispone de un período de prueba de {number_ofdays} días. \nAdemás disfruta de una garantía completa de {warranty}.\nDespachamos Envíos en 24h'
        elif warranty == None: # return, no warranty
            short_des = f'Este artículo dispone de un período de prueba de {number_ofdays} días.\nDespachamos Envíos en 24h'
    else: #no return days
        if warranty != None: #no returns, yes warranty
            short_des = f'Este artículo de una garantía completa de {warranty}.\nDespachamos Envíos en 24h'
        elif warranty == None:  #no returns, no warranty
            short_des = f'Despachamos Envíos en 24h'            

    return short_des


#write to filter_output_t2, sheet1 if data is ok, sheet2 = data to review by human
def recordto_t2(sheet, data_torecord):
    from openpyxl import load_workbook

    wb = load_workbook(FILTER_T2_OUTPUT)
    ws = wb[sheet]

    query   = data_torecord.get('query')
    target_prod_state   = data_torecord.get('target_prod_state')
    ebay_price  = data_torecord.get('ebay_price')
    ebay_shipping_time  = data_torecord.get('ebay_shipping_time')
    ebay_prod_specs  = data_torecord.get('ebay_prod_specs')
    ebay_prod_id    = data_torecord.get('ebay_prod_id')
    target_category = data_torecord.get('target_category')
    target_attr_1   = data_torecord.get('target_attr_1')
    target_attr_2   = data_torecord.get('target_attr_2')
    ebay_total_price    = data_torecord.get('ebay_total_price')
    ebay_price  = data_torecord.get('ebay_price')
    ebay_shipping_price = data_torecord.get('ebay_shipping_price')
    ebay_returns    = data_torecord.get('ebay_returns')
    ebay_prod_url   = data_torecord.get('ebay_prod_url')
    wp_price    = data_torecord.get('wp_price')
    detected_color  = data_torecord.get('detected_color')
    warranty    = data_torecord.get('warranty')
    target_model    = data_torecord.get('target_model')
    ads_pics        = data_torecord.get('ads_pics')
    wp_ebay_title   = data_torecord.get('ebay_title')
    wp_short_description = data_torecord.get('wp_short_description')
    ebay_pics   = data_torecord.get('ebay_pics')
    web_pics    = data_torecord.get('web_pics')
    web_pics = str(web_pics)

    last_row_s1 = ws.max_row + 1

    # if they exists, unpack ads_pics into 2 variable strings, easier to handle later
    #now I record all in the same cell, ad_pics is a string with all pics separated by comma
    if ads_pics:
        print(f'these are ads_pics {ads_pics}')
        # ad1 = str(ads_pics[0])
        # ad2 = str(ads_pics[1])
        # ws.cell(row=last_row_s1, column= ADS_PICS2_1_COL,value=  ad1)
        # ws.cell(row=last_row_s1, column= ADS_PICS2_2_COL,value=  ad2)
        string_pics = ','.join(ads_pics)
        ws.cell(row=last_row_s1, column= ADS_PICS2_1_COL,value=  string_pics)

    ws.cell(row=last_row_s1, column= TARGET_CATEGORY2_COL,value=  target_category)
    ws.cell(row=last_row_s1, column= COMPLETE_QUERY2_COL,value=  query)
    ws.cell(row=last_row_s1, column= WP_TITLE2_COL,value=  wp_ebay_title)
    # ws.cell(row=last_row_s1, column= WOO_ID2_COL,value=  '') #woo_id not available until wp_importer.py uploads
    ws.cell(row=last_row_s1, column= TARGET_STATE2_COL,value=  target_prod_state)
    # ws.cell(row=last_row_s1, column= SOURCE_STATE2_COL,value=  query) #source state not needed
    ws.cell(row=last_row_s1, column= ATTR1_1_COL,value=  target_attr_1)
    ws.cell(row=last_row_s1, column= ATTR1_2_COL,value=  target_attr_2)
    ws.cell(row=last_row_s1, column= WP_PRICE2_COL,value=  wp_price)
    # ws.cell(row=last_row_s1, column= WP_URL2_COL,value=  '') #not available until wp_importer.py uploads
    ws.cell(row=last_row_s1, column= SUPPLIER_URL2_COL,value=  ebay_prod_url)
    ws.cell(row=last_row_s1, column= EBAY_SHIPPING_TIME2_COL,value=  ebay_shipping_time)
    ws.cell(row=last_row_s1, column= EBAY_SHIPPING_PRICE2_COL,value=  ebay_shipping_price)
    ws.cell(row=last_row_s1, column= WARRANTY2_COL, value=  warranty)
    ws.cell(row=last_row_s1, column= EBAY_RETURNS2_COL,value=  ebay_returns)
    ws.cell(row=last_row_s1, column= WEB_PICS2_COL,value=  web_pics)
    ws.cell(row=last_row_s1, column= EBAY_ID2_COL,value=  ebay_prod_id)
    ws.cell(row=last_row_s1, column= TARGET_MODEL2_COL,value=  target_model)
    ws.cell(row=last_row_s1, column= EBAY_PICS2_COL,value=  ebay_pics)
    ws.cell(row=last_row_s1, column= SOURCE_SPECS, value=  ebay_prod_specs)
    ws.cell(row=last_row_s1, column= WP_SHORT_DESCRIPTION_COL, value=  wp_short_description)
    # ws.cell(row=last_row_s1, column= VARIABLE_PROD2_COL,value=  ) #variable prods ignored for now

    wb.save(FILTER_T2_OUTPUT)
    # print('SAVED FILE')


#read filter t1 output, search for column with checkmark
#compare target_prod_state == ebay_prod_state, if not, ebay state wins
#put the web pics id's
    # search for item model + attr1
    # manually check if they're the right color, mark that prod as reviewed, sheet3
    # write to an excel, sheet 2
#same for ad pics
#move file to logs
#clean file for the next round


#READ FILTER t1 OUTPUT, look for rows with checkmark
def run():
    from openpyxl import load_workbook

    wb = load_workbook(FILTER_OUTPUT1)
    ws = wb.active

    #clean f2 output before sending new data
    clean_excel(FILTER_T2_OUTPUT)

    current_row = START_ROW
    total_rows = len(ws['A'])
    print(total_rows)
    for row in range(total_rows):
        try:
            print(f'processed row {row}')

            query =  ws.cell(row=current_row, column= QUERY_COL).value
            # print(query)

            #ignore rows without checkmark
            check_mark =  ws.cell(row=current_row, column= CHECKMARK).value
            if check_mark == None : 
                current_row += 1
                continue 

            print('----------------------')

            query =         ws.cell(row=current_row, column= QUERY_COL).value
            ebay_title =    ws.cell(row=current_row, column= EBAY_TITLE_COL).value
            ebay_prod_state =    ws.cell(row=current_row, column= EBAY_PROD_STATE_COL).value
            target_prod_state =     ws.cell(row=current_row, column= TARGET_PROD_STATE_COL).value
            ebay_price =    ws.cell(row=current_row, column= EBAY_PRICE_COL).value
            ebay_shipping_time =    ws.cell(row=current_row, column= EBAY_SHIPPING_TIME_COL).value
            ebay_prod_id =  ws.cell(row=current_row, column= EBAY_PROD_ID_COL).value
            # ebay_category = ws.cell(row=current_row, column= EBAY_CATEGORY_COL).value
            ebay_prod_description =  ws.cell(row=current_row, column= EBAY_PROD_DESCRIPTION_COL).value
            ebay_pics =         ws.cell(row=current_row, column= PICTURES_COL).value
            target_category =   ws.cell(row=current_row, column= TARGET_CATEGORY_COL).value
            target_attr_2 =     ws.cell(row=current_row, column= DETECTED_COLOR_COL).value
            ebay_price =        ws.cell(row=current_row, column= EBAY_PRICE_COL).value
            ebay_shipping_price =   ws.cell(row=current_row, column= EBAY_SHIPPING_PRICE_COL).value
            ebay_returns =      ws.cell(row=current_row, column= EBAY_RETURNS_COL).value
            ebay_prod_url =     ws.cell(row=current_row, column= EBAY_PROD_URL_COL).value
            ebay_prod_specs =   ws.cell(row=current_row, column= EBAY_PROD_SPECS_COL).value
            wp_price =          ws.cell(row=current_row, column= WP_PRICE_COL).value
            # wp_shipping_time =  ws.cell(row=current_row, column= WP_SHIPPING_TIME_COL).value
            target_attr_1 =     ws.cell(row=current_row, column= TARGET_ATTR_1_COL).value
            ebay_total_price =  ws.cell(row=current_row, column= EBAY_TOTAL_PRICE_COL).value
            # ebay_vendor_notes =     ws.cell(row=current_row, column= EBAY_VENDOR_NOTES_COL).value
            # seller_votes =   ws.cell(row=current_row, column= EBAY_SELLER_VOTES_COL).value
            detected_color = ws.cell(row=current_row, column= DETECTED_COLOR_COL ).value
            warranty =       ws.cell(row=current_row, column= DETECTED_WARRANTY_COL ).value
            # subtitle =       ws.cell(row=current_row, column= SUBTITLE_COL).value
            available_colors =  ws.cell(row=current_row, column= AVAILABLE_COLORS_COL).value
            available_colors = str(available_colors) # commas makes it tupple
            target_model =   ws.cell(row=current_row, column= MODEL_COL).value
            target_model = str(target_model).lower()
            
            print(f'target_model: {target_model},target_attr_1: {target_attr_1} target_attr_2 {target_attr_2}') 

            current_row += 1 #to get to the next row in the next iteration
            
            prod_state = compare_states(target_prod_state, ebay_prod_state)
            
            # this makes that all pics used in web are ebay pics and not the saved ones
            web_pics = ebay_pics
            # web_pics =  get_pics_ids('web_pics', target_model, target_attr_2)
            # print(f'web_pics: {web_pics}')
            # if 'None' in web_pics or web_pics == 'not found':
            #     #check if ebay_pics are normal, not like 's-l500.jpg,s-l500.jpg,s-....'
            #     #if ebay_pics are ok, use them as web_pics
            #     #else, continue, ignore that prod

            #     #
            #     if 's-l500.jpg,s-l500.jpg,' not in ebay_pics:
            #         web_pics = ebay_pics
            #     else:
            #         print(f'not web pics for this source prod {ebay_prod_url}')
            #         continue
            # else:
            #     print(f'web_pics {web_pics}')
            
            ads_pics = get_pics_ids('ads_pics',target_model, target_attr_2)

            wp_short_description = get_short_description(ebay_returns, warranty)

            #need to know woo_id before upload to prods_db. 

            #create FILTER_T2_OUTPUT.xlsx
            # sheet1 = prods to upload adn get woo_id
            # sheet2 = prods with something missing, pics, warranty, etc...        

            data = {
                'query':query,
                'target_prod_state':target_prod_state,
                'ebay_price':ebay_price,
                'ebay_shipping_time':ebay_shipping_time,
                'ebay_prod_id':ebay_prod_id,
                'target_category':target_category,
                'target_attr_1':target_attr_1,
                'target_attr_2':target_attr_2,
                'ebay_total_price':ebay_total_price,
                'ebay_price':ebay_price,
                'ebay_shipping_price':ebay_shipping_price,
                'ebay_returns':ebay_returns,
                'ebay_prod_url':ebay_prod_url,
                'wp_price':wp_price,
                'detected_color':detected_color,
                'warranty':warranty,
                'target_model':target_model,
                'web_pics':web_pics,
                'ads_pics':ads_pics,
                'ebay_title':ebay_title,
                'ebay_pics':ebay_pics,
                'ebay_prod_specs':ebay_prod_specs,
                'wp_short_description':wp_short_description
                #woo_id is included after wp_importer.py
                }        

            #record to FILTER_T2_OUTPUT 
            recordto_t2('Sheet1', data)
            ## old, this recorded to sheet1 if all ok and sheet2 if some issue
            #if some issue, record to sheet2, if is all right record to sheet1
            # if ads_pics == 'not found' or ads_pics == None or web_pics == 'not found' or web_pics == None:
            #     recordto_t2('Sheet2', data)
            #     print('not found pics = written data to Sheet2')
            # else:
            #     recordto_t2('Sheet1', data)

            #output FILTER_T2_OUTPUT sheet wp
    
        except Exception as e:
            print(f'error in run(): {e}')
            traceback.print_exc()


if __name__ == '__main__':
    run()


#TRANSLATE
# deepl_auth_key = 'xxx'
# import deepl
# text = 'Der Endverbraucher ist verpflichtet Batterien ordnungsgemäß bei keiner Verwendung mehr zu entsorgen. Denn Batterien gehören nicht in den Hausmüll! Gerne helfen wir Ihnen bei Informationen zu Standorten in Ihrer Nähe. Detaillierte Informationen finden Sie in unseren AGB."'
# target_language = 'ES'
# translator = deepl.Translator(deepl_auth_key) 
# result = translator.translate_text(text, target_lang=target_language) 
# translated_text = result.text
# print(translated_text)

#EBAY API CONNECTION
# from ebaysdk.finding import Connection
# from ebaysdk.shopping import Connection as Shopping

# api = Connection(appid=XXXXX, config_file=None)
# api = Connection(config_file='ebay.yaml')
# response = api.execute('findItemsAdvanced', {'keywords': 'legos'})
# print(response)


# ebay_prod_id = 114974576186
# # api = Shopping(config_file='ebay.yaml')
# api = Shopping(appid='XXX', config_file=None)
# request = {'ItemID':ebay_prod_id}
# response = api.execute('GetSingleItem', request)
# print(response)

# api = Shopping(config_file='ebay.yaml')
# response = api.execute('FindPopularItems', {'QueryKeywords': 'Python'})
# print(response.dict())


# id="icImg"
# class="img img500 vi-img-gallery-vertical-align "

# from requests_html import HTMLSession
# session = HTMLSession()
# r = session.get('https://i.ebayimg.com/images/g/70QAAOSwz29hHOKL/')
# print(r.text)